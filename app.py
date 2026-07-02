# ═══════════════════════════════════════════════════════════════════════
#  ERB Manager — Rapprochement Bancaire  (Streamlit)
#  v5.6 — Corrections : soldes persistants + téléchargement + clôture
# ═══════════════════════════════════════════════════════════════════════
import sys, asyncio
import random, time
if sys.platform == "win32" and sys.version_info >= (3, 12):
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
import re
import numpy as np
import pandas as pd
import streamlit as st
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def _extract_pdf_tables(uploaded_file):
    """
    Extrait les données d'un PDF bancaire.
    Stratégies (dans l'ordre) :
      1. pdfplumber tables structurées
      2. pdfplumber texte brut → parsing intelligent
      3. Détection scan → message explicite
    """
    try:
        import pdfplumber, io, re as _re
        uploaded_file.seek(0)
        raw_bytes = uploaded_file.read()
        all_rows = []
        n_pages = 0
        has_text = False

        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            n_pages = len(pdf.pages)
            for page_num, page in enumerate(pdf.pages):

                # ── Stratégie 1 : tableaux structurés ────────────────
                tables = page.extract_tables({
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                })
                if not tables:
                    tables = page.extract_tables({
                        "vertical_strategy": "text",
                        "horizontal_strategy": "text",
                        "snap_tolerance": 5,
                    })
                for table in tables:
                    for row in table:
                        if row and any(c for c in row if c and str(c).strip()):
                            all_rows.append([str(c or '').strip().replace('\n',' ') for c in row])

                # ── Stratégie 2 : texte brut si pas de tableaux ───────
                if not tables:
                    text = page.extract_text(x_tolerance=3, y_tolerance=3)
                    if text and text.strip():
                        has_text = True
                        for line in text.split('\n'):
                            line = line.strip()
                            if not line: continue
                            # Détecter lignes avec dates (format africain DD/MM/YYYY ou DD-MM-YYYY)
                            # Séparer sur multiples espaces pour préserver les champs
                            parts = _re.split(r'  +', line)
                            if len(parts) >= 2:
                                all_rows.append([p.strip() for p in parts])
                            else:
                                # Tentative de split sur tabulation
                                parts2 = line.split('\t')
                                if len(parts2) >= 2:
                                    all_rows.append([p.strip() for p in parts2])
                                else:
                                    all_rows.append([line])

        if not all_rows:
            if not has_text:
                return None, (
                    f"PDF scanné détecté ({n_pages} page(s)) — ce type de PDF contient des images "
                    "et non du texte. Solutions : \n"
                    "• Demandez le relevé en format numérique (export PDF depuis le portail bancaire)\n"
                    "• Ou convertissez manuellement en Excel avant import"
                )
            return None, f"Aucune donnée extractible trouvée dans le PDF ({n_pages} page(s))"

        # Normaliser les longueurs
        max_cols = max(len(r) for r in all_rows)
        all_rows = [r + [''] * (max_cols - len(r)) for r in all_rows]
        df = pd.DataFrame(all_rows)
        return df, None

    except ImportError:
        return None, "pdfplumber non installé. Ajoutez 'pdfplumber' dans requirements.txt"
    except Exception as ex:
        return None, f"Erreur lecture PDF : {str(ex)[:300]}"


# ── Configuration Supabase ────────────────────────────────────────────
# 1. Créez un projet sur supabase.com (gratuit)
# 2. Allez dans Settings → API → copiez l'URL et la clé anon
# 3. Remplacez ci-dessous OU ajoutez dans Streamlit Secrets :
#    SUPABASE_URL = "https://xxxx.supabase.co"
#    SUPABASE_KEY = "eyJ..."
SUPABASE_URL = ""
SUPABASE_KEY = ""

def get_supabase():
    """Retourne un client Supabase configuré (depuis secrets ou constantes)."""
    try:
        from supabase import create_client
        # Essayer st.secrets, mais ignorer si fichier absent
        url, key = SUPABASE_URL, SUPABASE_KEY
        try:
            url = st.secrets.get("SUPABASE_URL", SUPABASE_URL)
            key = st.secrets.get("SUPABASE_KEY", SUPABASE_KEY)
        except Exception:
            pass  # Pas de secrets.toml → utiliser les constantes ci-dessus
        if "VOTRE_PROJET" in url or "VOTRE_CLE" in key or not url or not key:
            return None
        return create_client(url, key)
    except Exception as _e:
        st.session_state["_supabase_error"] = str(_e)
        return None

def sauvegarder_erb(entreprise, licence_code, periode, xlsx_bytes, sr_rb, sr_cp, ok):
    """
    Sauvegarde un ERB dans Supabase.
    Table requise (à créer dans Supabase SQL Editor) :
    
    CREATE TABLE erb_historique (
        id            BIGSERIAL PRIMARY KEY,
        entreprise    TEXT NOT NULL,
        licence_code  TEXT NOT NULL,
        periode       TEXT NOT NULL,
        date_creation TIMESTAMPTZ DEFAULT NOW(),
        sr_rb         NUMERIC,
        sr_cp         NUMERIC,
        equilibre     BOOLEAN,
        xlsx_base64   TEXT
    );
    """
    sb = get_supabase()
    if sb is None:
        return False, "Supabase non configuré"
    try:
        import base64
        xlsx_b64 = base64.b64encode(xlsx_bytes).decode('utf-8')
        data = {
            "entreprise":   entreprise,
            "licence_code": licence_code,
            "periode":      periode,
            "sr_rb":        float(sr_rb),
            "sr_cp":        float(sr_cp),
            "equilibre":    bool(ok),
            "xlsx_base64":  xlsx_b64,
        }
        result = sb.table("erb_historique").insert(data).execute()
        return True, ""
    except Exception as ex:
        return False, str(ex)

def charger_historique_erb(licence_code, limit=20):
    """Charge les N derniers ERB sauvegardés pour ce client."""
    sb = get_supabase()
    if sb is None:
        return []
    try:
        result = (sb.table("erb_historique")
                    .select("id,entreprise,periode,date_creation,sr_rb,sr_cp,equilibre")
                    .eq("licence_code", licence_code)
                    .order("date_creation", desc=True)
                    .limit(limit)
                    .execute())
        return result.data or []
    except Exception:
        return []

def telecharger_erb_sauvegarde(erb_id):
    """Récupère les bytes Excel d'un ERB sauvegardé par son ID."""
    sb = get_supabase()
    if sb is None:
        return None
    try:
        import base64
        result = (sb.table("erb_historique")
                    .select("xlsx_base64,periode,entreprise")
                    .eq("id", erb_id)
                    .single()
                    .execute())
        if result.data:
            xlsx_bytes = base64.b64decode(result.data["xlsx_base64"])
            fname = f"ERB_{result.data['entreprise']}_{result.data['periode'].replace(' ','_')}.xlsx"
            return xlsx_bytes, fname
        return None, None
    except Exception:
        return None, None


st.set_page_config(page_title="ERB Manager — Rapprochement Bancaire", page_icon="🏦", layout="wide")

# ── CSS premium — injecté une seule fois au démarrage ─────────────────
st.markdown("""<style>
/* ── Fond principal dégradé ── */
.stApp {
    background: linear-gradient(135deg, #0F1B2D 0%, #1A2E4A 40%, #0D2137 70%, #162032 100%) !important;
    background-attachment: fixed !important;
}
.main .block-container {
    background: rgba(255,255,255,0.04) !important;
    border-radius: 16px !important;
    border: 1px solid rgba(255,255,255,0.08) !important;
    backdrop-filter: blur(10px) !important;
    padding: 2rem 2.5rem !important;
    margin-top: 0.5rem !important;
}
.stApp, .stApp p, .stApp label, .stApp div { color: #E8EDF4 !important; }
.stApp h1, .stApp h2, .stApp h3 { color: #FFFFFF !important; }
/* ── Sidebar ── */
section[data-testid="stSidebar"] > div:first-child { padding-top: 1rem; }
section[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #1A237E 0%, #283593 30%, #1565C0 65%, #0D47A1 100%) !important;
    border-right: 1px solid rgba(100,160,255,0.2) !important;
}
section[data-testid="stSidebar"] * { color: #E3F2FD !important; }
section[data-testid="stSidebar"] hr { border-color: rgba(100,160,255,0.25) !important; }
section[data-testid="stSidebar"] button {
    background: rgba(255,255,255,0.08) !important;
    border: 1px solid rgba(100,160,255,0.25) !important;
    border-radius: 10px !important; color: #E3F2FD !important;
}
section[data-testid="stSidebar"] button:hover {
    background: rgba(100,160,255,0.2) !important;
    border-color: rgba(100,160,255,0.5) !important;
}
section[data-testid="stSidebar"] button[kind="primary"] {
    background: linear-gradient(135deg, #1976D2, #42A5F5) !important;
    border-color: #64B5F6 !important; font-weight: 700 !important;
    box-shadow: 0 4px 15px rgba(66,165,245,0.4) !important;
}
section[data-testid="stSidebar"] button p { color: white !important; }
/* ── Boutons principaux ── */
.stButton button[kind="primary"] {
    background: linear-gradient(135deg, #1565C0, #1976D2, #42A5F5) !important;
    border: none !important; border-radius: 10px !important; color: white !important;
    font-weight: 600 !important; box-shadow: 0 4px 20px rgba(66,165,245,0.4) !important;
}
.stButton button[kind="primary"]:hover { box-shadow: 0 6px 25px rgba(66,165,245,0.6) !important; transform: translateY(-1px) !important; }
.stButton button[kind="secondary"] {
    background: rgba(255,255,255,0.07) !important; border: 1px solid rgba(255,255,255,0.18) !important;
    border-radius: 10px !important; color: #B0C4DE !important;
}
.stButton button[kind="secondary"]:hover { background: rgba(255,255,255,0.13) !important; color: white !important; }
/* ── Inputs ── */
.stTextInput input, .stNumberInput input {
    background: rgba(10,25,50,0.7) !important; border: 1.5px solid rgba(100,160,255,0.4) !important;
    border-radius: 8px !important; color: #FFFFFF !important; font-weight: 500 !important; font-size: 14px !important;
}
.stTextInput input:focus, .stNumberInput input:focus {
    border-color: #42A5F5 !important; box-shadow: 0 0 0 3px rgba(66,165,245,0.25) !important;
    background: rgba(10,30,60,0.85) !important;
}
.stTextInput label, .stNumberInput label, .stSelectbox label {
    color: #90CAF9 !important; font-weight: 600 !important; font-size: 13px !important;
}
.stTextInput input::placeholder, .stNumberInput input::placeholder {
    color: rgba(144,202,249,0.5) !important; font-style: italic !important;
}
/* ── Métriques ── */
[data-testid="metric-container"] {
    background: linear-gradient(135deg, rgba(21,101,192,0.25), rgba(30,50,100,0.35)) !important;
    border: 1px solid rgba(100,160,255,0.2) !important; border-radius: 14px !important;
    padding: 1rem 1.2rem !important; box-shadow: 0 4px 20px rgba(0,0,0,0.3) !important;
}
[data-testid="metric-container"] label { color: #90CAF9 !important; font-size: 13px !important; font-weight: 600 !important; text-transform: uppercase !important; }
[data-testid="metric-container"] [data-testid="stMetricValue"] { color: #FFFFFF !important; font-size: 2rem !important; font-weight: 700 !important; }
/* ── Expander ── */
.streamlit-expanderHeader {
    background: rgba(255,255,255,0.06) !important; border: 1px solid rgba(100,160,255,0.18) !important;
    border-radius: 10px !important; color: #B0C4DE !important; font-weight: 600 !important;
}
.streamlit-expanderContent {
    background: rgba(255,255,255,0.03) !important; border: 1px solid rgba(100,160,255,0.12) !important;
    border-top: none !important; border-radius: 0 0 10px 10px !important;
}
/* ── Alerts ── */
.stSuccess { background: linear-gradient(135deg,rgba(27,94,32,0.6),rgba(46,125,50,0.4)) !important; border:1px solid #4CAF50 !important; border-radius:10px !important; color:#C8E6C9 !important; }
.stWarning { background: linear-gradient(135deg,rgba(230,81,0,0.4),rgba(245,124,0,0.3)) !important; border:1px solid #FF9800 !important; border-radius:10px !important; color:#FFE0B2 !important; }
.stInfo    { background: linear-gradient(135deg,rgba(13,71,161,0.5),rgba(21,101,192,0.35)) !important; border:1px solid #1976D2 !important; border-radius:10px !important; color:#BBDEFB !important; }
.stError   { background: linear-gradient(135deg,rgba(183,28,28,0.5),rgba(198,40,40,0.35)) !important; border:1px solid #EF5350 !important; border-radius:10px !important; }
hr { border-color: rgba(100,160,255,0.2) !important; }
/* ── Dataframe ── */
[data-testid="stDataFrame"] { border-radius: 10px !important; overflow: hidden !important; border: 1px solid rgba(100,160,255,0.2) !important; }
/* ── File uploader ── */
[data-testid="stFileUploader"] { background: rgba(255,255,255,0.04) !important; border: 2px dashed rgba(100,160,255,0.3) !important; border-radius: 12px !important; }
/* ── Download button ── */
[data-testid="stDownloadButton"] button { background: linear-gradient(135deg,#1B5E20,#2E7D32,#388E3C) !important; border:none !important; border-radius:10px !important; color:white !important; font-weight:600 !important; box-shadow:0 4px 15px rgba(56,142,60,0.4) !important; }
/* ── Scrollbar ── */
::-webkit-scrollbar { width:6px; height:6px; }
::-webkit-scrollbar-track { background: rgba(255,255,255,0.03); }
::-webkit-scrollbar-thumb { background: rgba(100,160,255,0.3); border-radius:3px; }
/* ── Classes custom ── */
.page-title { font-size:26px; font-weight:800; color:#FFFFFF; margin-bottom:4px; }
.page-sub   { font-size:13px; color:#78909C; margin-bottom:1.4rem; }
.logo-box   { background:linear-gradient(135deg,#1D9E75,#26A69A); border-radius:12px; padding:10px 14px; color:white; font-weight:800; font-size:16px; margin-bottom:6px; box-shadow:0 4px 15px rgba(29,158,117,0.5); }
.logo-sub   { font-size:11px; color:rgba(163,210,230,0.85); margin-bottom:1rem; padding-left:2px; }
</style>""", unsafe_allow_html=True)

# ═══════════════════════════════════════════════════════════════════════
#  SYSTÈME DE LICENCES — ERB Manager
#  OTP = code à 6 chiffres envoyé par email, valable 10 minutes
#  Pour ajouter un client : ajoutez une entrée dans LICENCES
#  Pour désactiver : passez "actif" à False
# ═══════════════════════════════════════════════════════════════════════

# ── Configuration SendGrid ────────────────────────────────────────────
# Créez un compte sur sendgrid.com, générez une API Key,
# et remplacez la valeur ci-dessous (ou utilisez st.secrets)
SENDGRID_API_KEY  = ""
EMAIL_EXPEDITEUR  = "sowmarieta013@gmail.com"   # email vérifié dans SendGrid
NOM_EXPEDITEUR    = "ERB Manager"

LICENCES = {
    # "CODE": {
    #   "entreprise": "Nom affiché",
    #   "domaines":   ["@entreprise.com"],  # domaines email autorisés
    #   "actif":      True/False
    # }
    "ERB-4C11-316D-54AB": {
        "entreprise": "Alliance — Amadou Lamine MBODJ",
        "domaines":   ["@Alliance-ac.sn"],
        "actif": True,
    },
    "ERB-3D3A-9BC8-037B": {
        "entreprise": "Alliance — Khardiatou DIALLO KA",
        "domaines":   ["@Alliance-ac.sn"],
        "actif": True,
    },
    "ERB-AA5D-5BA6-58B1": {
        "entreprise": "Alliance — Mamadou DIALLO",
        "domaines":   ["@Alliance-ac.sn"],
        "actif": True,
    },
    "ERB-1FF5-8FBA-FC26": {
        "entreprise": "Client 4",
        "domaines":   ["@gmail.com"],
        "actif": True,
    },
    "ERB-B987-B122-E9BE": {
        "entreprise": "Client 5",
        "domaines":   ["@gmail.com"],
        "actif": True,
    
    },
}


def _email_domain_ok(email, domaines):
    """Vérifie que l'email appartient à un domaine autorisé."""
    return any(email.strip().lower().endswith(d.lower()) for d in domaines)

def _send_otp(email, otp, entreprise):
    """
    Envoie le code OTP par email.
    Essaie dans l'ordre :
      1. Resend.com  (si RESEND_API_KEY configuré — recommandé sur Streamlit Cloud)
      2. SendGrid    (si SENDGRID_API_KEY configuré)
    Retourne (True, '') ou (False, message_erreur).
    """
    import ssl, urllib.request, json as _json

    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode    = ssl.CERT_NONE

    html_body = f"""<div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;
        background:#0F1B2D;border-radius:12px;padding:32px;color:#E8EDF4">
      <div style="text-align:center;margin-bottom:24px">
        <div style="font-size:22px;font-weight:800;color:#fff">ERB Manager</div>
      </div>
      <p style="color:#90CAF9">Bonjour,</p>
      <p style="color:#B0C4DE">Voici votre code de connexion pour <strong style="color:#fff">{entreprise}</strong> :</p>
      <div style="background:#1A2E4A;border:1px solid #42A5F5;border-radius:12px;
          text-align:center;padding:20px;margin:24px 0">
        <div style="font-size:40px;font-weight:900;letter-spacing:12px;color:#42A5F5;font-family:monospace">{otp}</div>
        <div style="font-size:12px;color:#78909C;margin-top:8px">Valable 10 minutes</div>
      </div>
      <p style="font-size:12px;color:#546E7A">Ne partagez jamais ce code.</p>
    </div>"""

    # ── Essai 1 : Resend.com ─────────────────────────────────────────
    resend_key = ""
    try:
        resend_key = st.secrets.get("RESEND_API_KEY", "")
    except Exception:
        pass  # Pas de secrets.toml en local

    if resend_key and not resend_key.startswith("re.VOTRE"):
        try:
            payload = _json.dumps({
                "from":    f"{NOM_EXPEDITEUR} <{EMAIL_EXPEDITEUR}>",
                "to":      [email],
                "subject": f"[ERB Manager] Votre code : {otp}",
                "html":    html_body,
            }).encode("utf-8")
            req = urllib.request.Request(
                "https://api.resend.com/emails",
                data=payload,
                headers={"Authorization": f"Bearer {resend_key}", "Content-Type": "application/json"},
                method="POST"
            )
            with urllib.request.urlopen(req, context=ctx, timeout=15) as r:
                if r.status in (200, 201):
                    return True, ""
        except urllib.error.HTTPError as e:
            resend_err = f"Resend HTTP {e.code}: {e.read().decode('utf-8','ignore')[:150]}"
        except Exception as ex:
            resend_err = str(ex)

    # ── Essai 2 : SendGrid ───────────────────────────────────────────
    api_key = SENDGRID_API_KEY
    try:
        api_key = st.secrets.get("SENDGRID_API_KEY", api_key)
    except Exception:
        pass  # Pas de secrets.toml en local

    if api_key.startswith("SG.VOTRE"):
        return False, "Aucun service email configuré (SendGrid ou Resend). Voir la configuration."

    html_body = f"""
    <div style="font-family:Arial,sans-serif;max-width:480px;margin:0 auto;
                background:#0F1B2D;border-radius:12px;padding:32px;color:#E8EDF4">
      <div style="text-align:center;margin-bottom:24px">
        <div style="font-size:40px">&#127974;</div>
        <div style="font-size:22px;font-weight:800;color:#fff">ERB Manager</div>
      </div>
      <p style="color:#90CAF9;margin-bottom:8px">Bonjour,</p>
      <p style="color:#B0C4DE">Voici votre code de connexion pour
         <strong style="color:#fff">{entreprise}</strong> :</p>
      <div style="background:#1A2E4A;border:1px solid #42A5F5;border-radius:12px;
                  text-align:center;padding:20px;margin:24px 0">
        <div style="font-size:40px;font-weight:900;letter-spacing:12px;
                     color:#42A5F5;font-family:monospace">{otp}</div>
        <div style="font-size:12px;color:#78909C;margin-top:8px">
          Valable 10 minutes — usage unique
        </div>
      </div>
      <p style="font-size:12px;color:#546E7A">
        Si vous n'avez pas demandé ce code, ignorez cet email.<br>
        Ne partagez jamais ce code.
      </p>
      <hr style="border-color:#1A2E4A;margin:24px 0">
      <p style="font-size:11px;color:#37474F;text-align:center">
        ERB Manager — Rapprochement Bancaire Professionnel
      </p>
    </div>"""

    payload = _json.dumps({
        "personalizations": [{"to": [{"email": email}]}],
        "from": {"email": EMAIL_EXPEDITEUR, "name": NOM_EXPEDITEUR},
        "subject": f"[ERB Manager] Votre code : {otp}",
        "content": [{"type": "text/html", "value": html_body}]
    }).encode("utf-8")

    payload = _json.dumps({
        "personalizations": [{"to": [{"email": email}]}],
        "from": {"email": EMAIL_EXPEDITEUR, "name": NOM_EXPEDITEUR},
        "subject": f"[ERB Manager] Votre code : {otp}",
        "content": [{"type": "text/html", "value": html_body}]
    }).encode("utf-8")

    try:
        req = urllib.request.Request(
            "https://api.sendgrid.com/v3/mail/send",
            data    = payload,
            headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
            method  = "POST",
        )
        with urllib.request.urlopen(req, context=ctx, timeout=15) as resp:
            if resp.status in (200, 202):
                return True, ""
            return False, f"SendGrid statut {resp.status}"
    except urllib.error.HTTPError as e:
        body = e.read().decode("utf-8", errors="ignore")
        # 403 = domaine bloqué sur Streamlit Cloud → erreur explicite
        if e.code == 403 and "allowlist" in body:
            return False, "SendGrid bloqué par Streamlit Cloud. Utilisez Resend.com à la place (ajoutez RESEND_API_KEY dans les Secrets)."
        return False, f"SendGrid HTTP {e.code} : {body[:200]}"
    except Exception as ex:
        return False, str(ex)

def check_licence():
    """
    Connexion en 3 étapes :
      1. Code de licence  → identifie l'entreprise + domaines autorisés
      2. Email professionnel → vérifié contre les domaines + envoi OTP
      3. Code OTP → validé (6 chiffres, 10 min, usage unique)
    """
    if st.session_state.get("_licence_ok"):
        return True

    # CSS page connexion
    st.markdown("""<style>
    .stApp { background: linear-gradient(135deg,#0F1B2D,#1A2E4A,#0D2137) !important; }
    .main .block-container { max-width:500px !important; margin:2.5rem auto !important;
        background:rgba(255,255,255,0.05) !important; border-radius:20px !important;
        border:1px solid rgba(100,160,255,0.15) !important; padding:2.5rem 2.5rem !important; }
    </style>""", unsafe_allow_html=True)

    # En-tête
    st.markdown("""
    <div style="text-align:center;margin-bottom:1.8rem">
      <div style="font-size:48px;margin-bottom:8px">🏦</div>
      <div style="font-size:26px;font-weight:800;color:#fff">ERB Manager</div>
      <div style="font-size:13px;color:#78909C;margin-top:4px">Rapprochement Bancaire Professionnel</div>
    </div>""", unsafe_allow_html=True)

    step = st.session_state.get("_auth_step", 1)

    # Barre de progression
    pct = {1: "33%", 2: "66%", 3: "100%"}.get(step, "33%")
    labels = {1: "Étape 1/3 — Licence", 2: "Étape 2/3 — Email", 3: "Étape 3/3 — Vérification"}
    st.markdown(f"""
    <div style="margin-bottom:1.5rem">
      <div style="display:flex;justify-content:space-between;margin-bottom:6px">
        <span style="font-size:12px;color:#90CAF9;font-weight:600">{labels[step]}</span>
      </div>
      <div style="background:rgba(255,255,255,0.08);border-radius:4px;height:4px">
        <div style="background:linear-gradient(90deg,#1565C0,#42A5F5);width:{pct};
                    height:4px;border-radius:4px;transition:width 0.3s"></div>
      </div>
    </div>""", unsafe_allow_html=True)

    # ── ÉTAPE 1 : Code de licence ─────────────────────────────────────
    if step == 1:
        code = st.text_input("🔑 Code de licence", placeholder="ERB-XXXX-XXXX-XXXX",
                             help="Code fourni par ERB Manager lors de votre abonnement.")
        if st.button("→ Continuer", type="primary", use_container_width=True):
            code_clean = code.strip().upper()
            if code_clean in LICENCES and LICENCES[code_clean]["actif"]:
                st.session_state["_auth_step"]      = 2
                st.session_state["_auth_code_tmp"]  = code_clean
                st.session_state["_auth_info_tmp"]  = LICENCES[code_clean]
                st.rerun()
            elif code_clean in LICENCES and not LICENCES[code_clean]["actif"]:
                st.error("❌ Licence désactivée. Contactez ERB Manager.")
            else:
                st.error("❌ Code invalide. Vérifiez le code reçu.")

    # ── ÉTAPE 2 : Email professionnel ─────────────────────────────────
    elif step == 2:
        info       = st.session_state["_auth_info_tmp"]
        entreprise = info["entreprise"]
        domaines   = info["domaines"]
        dom_str    = "  ·  ".join(domaines)

        st.markdown(f"""
        <div style="background:rgba(66,165,245,0.1);border:1px solid rgba(66,165,245,0.25);
                    border-radius:10px;padding:10px 14px;margin-bottom:1.2rem">
          <div style="font-size:11px;color:#90CAF9;font-weight:700">✅ LICENCE RECONNUE</div>
          <div style="font-size:15px;color:#fff;font-weight:800">🏢 {entreprise}</div>
          <div style="font-size:11px;color:#607D8B;margin-top:2px">Domaines : {dom_str}</div>
        </div>""", unsafe_allow_html=True)

        email = st.text_input("📧 Votre email professionnel",
                              placeholder=f"prenom.nom{domaines[0]}",
                              help=f"Doit appartenir au domaine : {dom_str}")

        c1, c2 = st.columns([1, 2])
        with c1:
            if st.button("← Retour", use_container_width=True):
                for k in ["_auth_step","_auth_code_tmp","_auth_info_tmp"]:
                    st.session_state.pop(k, None)
                st.rerun()
        with c2:
            if st.button("📩 Recevoir le code", type="primary", use_container_width=True):
                email_clean = email.strip().lower()
                if "@" not in email_clean or "." not in email_clean.split("@")[-1]:
                    st.error("❌ Format email invalide.")
                elif not _email_domain_ok(email_clean, domaines):
                    st.error(f"❌ Email non autorisé pour {entreprise}. Domaines acceptés : {dom_str}")
                else:
                    # Générer OTP 6 chiffres
                    otp = str(random.randint(100000, 999999))
                    with st.spinner("📩 Envoi du code en cours..."):
                        ok, err = _send_otp(email_clean, otp, entreprise)
                    # Dans tous les cas, passer à l'étape 3
                    st.session_state["_auth_step"]     = 3
                    st.session_state["_auth_email"]    = email_clean
                    st.session_state["_auth_otp"]      = otp
                    st.session_state["_auth_otp_time"] = time.time()
                    # Toujours stocker le code (affiché si email non reçu)
                    st.session_state["_auth_debug_otp"] = otp
                    if not ok:
                        st.session_state["_auth_send_err"] = err
                    else:
                        st.session_state["_auth_send_err"] = ""
                    st.rerun()

    # ── ÉTAPE 3 : Vérification OTP ────────────────────────────────────
    elif step == 3:
        email      = st.session_state["_auth_email"]
        info       = st.session_state["_auth_info_tmp"]
        entreprise = info["entreprise"]
        otp_time   = st.session_state["_auth_otp_time"]
        remaining  = int(600 - (time.time() - otp_time))

        if remaining <= 0:
            st.warning("⏱️ Code expiré. Recommencez.")
            for k in ["_auth_step","_auth_code_tmp","_auth_info_tmp",
                      "_auth_email","_auth_otp","_auth_otp_time"]:
                st.session_state.pop(k, None)
            st.rerun()

        st.markdown(f"""
        <div style="background:rgba(29,158,117,0.1);border:1px solid rgba(29,158,117,0.25);
                    border-radius:10px;padding:10px 14px;margin-bottom:1.2rem">
          <div style="font-size:12px;color:#80CBC4;font-weight:700">📩 CODE ENVOYÉ À</div>
          <div style="font-size:14px;color:#fff;font-weight:700">{email}</div>
          <div style="font-size:11px;color:#607D8B;margin-top:2px">
            ⏱️ Expire dans {remaining // 60}m {remaining % 60:02d}s
          </div>
        </div>""", unsafe_allow_html=True)

        # Afficher le code à l'écran (fallback fiable)
        if st.session_state.get("_auth_debug_otp"):
            send_err = st.session_state.get("_auth_send_err","")
            if send_err:
                # Erreur d'envoi → affichage orange avec message d'erreur
                label = "⚠️ Email non reçu — Voici votre code :"
                border_color = "#FF9800"
                bg_color = "rgba(255,152,0,0.15)"
                note = f'<div style="font-size:11px;color:#888;margin-top:8px">Erreur : {send_err}</div>'
            else:
                # Email envoyé mais peut ne pas arriver → affichage bleu discret
                label = "🔐 Votre code de connexion :"
                border_color = "#42A5F5"
                bg_color = "rgba(66,165,245,0.1)"
                note = '<div style="font-size:11px;color:#90CAF9;margin-top:8px">Vérifiez aussi votre boîte email (et les spams)</div>'
            st.markdown(f"""
            <div style="background:{bg_color};border:1px solid {border_color};
                        border-radius:10px;padding:14px 18px;margin-bottom:1rem">
              <div style="font-size:12px;color:#FFB74D;font-weight:700;margin-bottom:6px">
                {label}
              </div>
              <div style="font-size:36px;font-weight:900;letter-spacing:10px;
                          color:#FFF;font-family:monospace;text-align:center;
                          background:rgba(0,0,0,0.3);border-radius:8px;padding:12px">
                {st.session_state["_auth_debug_otp"]}
              </div>
              {note}
            </div>""", unsafe_allow_html=True)

        otp_input = st.text_input("🔢 Code à 6 chiffres",
                                  placeholder="123456",
                                  max_chars=6,
                                  help="Vérifiez votre boîte email (et le dossier spam).")

        c1, c2 = st.columns([1, 2])
        with c1:
            if st.button("← Retour", use_container_width=True):
                for k in ["_auth_step","_auth_code_tmp","_auth_info_tmp",
                          "_auth_email","_auth_otp","_auth_otp_time"]:
                    st.session_state.pop(k, None)
                st.rerun()
        with c2:
            if st.button("✅ Se connecter", type="primary", use_container_width=True):
                if otp_input.strip() == st.session_state["_auth_otp"]:
                    # ✅ Accès accordé
                    st.session_state["_licence_ok"]   = True
                    st.session_state["_licence_code"] = st.session_state["_auth_code_tmp"]
                    st.session_state["entreprise"]    = entreprise
                    st.session_state["_user_email"]   = email
                    # Nettoyer toutes les clés temporaires d'auth
                    for k in ["_auth_step","_auth_code_tmp","_auth_info_tmp",
                              "_auth_email","_auth_otp","_auth_otp_time",
                              "_auth_debug_otp","_auth_send_err"]:
                        st.session_state.pop(k, None)
                    st.rerun()
                else:
                    st.error("❌ Code incorrect. Vérifiez le code reçu par email.")

        # Renvoi du code
        st.markdown('<div style="text-align:center;margin-top:1rem">', unsafe_allow_html=True)
        if st.button("🔄 Renvoyer le code", use_container_width=False):
            otp_new = str(random.randint(100000, 999999))
            with st.spinner("Envoi..."):
                ok, err = _send_otp(email, otp_new, entreprise)
            if ok:
                st.session_state["_auth_otp"]      = otp_new
                st.session_state["_auth_otp_time"] = time.time()
                st.success("📩 Nouveau code envoyé !")
                st.rerun()
            else:
                st.session_state["_auth_debug_otp"] = otp_new
                st.session_state["_auth_send_err"]  = err
                st.session_state["_auth_otp"]      = otp_new
                st.session_state["_auth_otp_time"] = time.time()
                st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("""
    <div style="text-align:center;margin-top:1.5rem;font-size:12px;color:#37474F">
      Problème de connexion ? <span style="color:#42A5F5;font-weight:600">contact@erbmanager.com</span>
    </div>""", unsafe_allow_html=True)
    return False


# ── Vérification licence ─────────────────────────────────────────────
if not check_licence():
    st.stop()

# ── Initialisation session_state ──────────────────────────────────────
DEFAULTS = dict(
    page="dashboard", rb_df=None, cp_df=None, rb_raw=None, cp_raw=None,
    rb_hd=None, cp_hd=None, rb_map={}, cp_map={},
    rb_solde_auto=None, cp_solde_auto=None, mc=0,
    carryover_rb=[], carryover_cp=[], historique=[],
    regularisees=[], rematch_needed=False,
    _rematching=False, _rematch_msg='',
    # Persistance des champs Informations entre reruns
    inf_cl='', inf_bq='', inf_co='', inf_pe='',
    s_rb=0.0, s_cp=0.0,
)
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ───────────────────────────────────────────────────────────────────────
# UTILITAIRES
# ───────────────────────────────────────────────────────────────────────
def pn(v):
    if v is None or (isinstance(v, float) and np.isnan(v)) or str(v).strip() in ('','nan','None','-','—'):
        return 0.0
    try: return float(str(v).replace(' ','').replace('\xa0','').replace(',','.'))
    except: return 0.0

def fmt_fr(v):
    if v is None or (isinstance(v, float) and np.isnan(v)): return ''
    n = float(v)
    if abs(n) < 0.005: return ''
    return f"{n:,.2f}".replace(',', ' ').replace('.', ',').replace(' ', '\u202f')

EXCLUDE_EXACT = {'totaux','total','totals'}
EXCLUDE_START = ['a.n. ','a.n.au','a-n au','a/n au','a.n au',
    'solde avant rapprochement','solde final','solde initial','solde au ',
    "solde de clôture","solde d'ouverture","balance d'ouverture",
    'opening balance','closing balance','report à nouveau','à nouveau au','report au']

def is_excluded(lib):
    l = str(lib).lower().strip()
    if l in EXCLUDE_EXACT: return True
    if any(l.startswith(k) for k in EXCLUDE_START): return True
    if re.match(r'^a\.n\.?\s', l) or l in ('a.n.','a.n'): return True
    return False

AM = {
    'date':    ["date de l'opération","date operation","date valeur","date comptable","date","dat","jour","date "],
    'lib':     ["libelle","libellé","libelle ","description","desc","motif","intitule","intitulé","wording"],
    'piece':   ["référence","reference","n° piece","n°piece","piece","ref","num","n°","voucher","n° pièce"],
    'debit':   ["debit","débit","debit ","deb","dbt","sortie","depense","montant debit"],
    'credit':  ["credit","crédit","credit ","crd","crt","entree","recette","montant credit"],
    'montant': ["montant","montant ","amount","valeur"],
    'sens':    ["sens","sens ","type","direction"],
}
FL = {'date':"Date",'lib':"Libellé",'piece':"N° Pièce",'debit':"Débit",'credit':"Crédit",
      'montant':"Montant (colonne unique)",'sens':"Sens (Débit/Crédit)"}

def guess_col(headers, field):
    keys = AM[field]
    for h in headers:
        if str(h).lower().strip() in keys: return h
    for h in headers:
        hn = str(h).lower().strip()
        for k in keys:
            if hn.startswith(k) and len(hn) <= len(k)+3: return h
    return None

def detect_header_row(df_raw):
    for i, row in df_raw.iterrows():
        if row.dropna().apply(lambda x: str(x).strip() != '').sum() >= 3:
            return i
    return 0

def _extract_pdf_scanned_via_claude(uploaded_file):
    """
    Utilise l'API Claude pour lire un relevé bancaire scanné (PDF image).
    Retourne un DataFrame avec les colonnes détectées.
    """
    try:
        import base64, json, io, urllib.request, ssl
        uploaded_file.seek(0)
        raw_bytes = uploaded_file.read()

        # Récupérer la clé API Anthropic
        api_key = ""
        try:
            api_key = st.secrets.get("ANTHROPIC_API_KEY", "")
        except Exception:
            pass

        if not api_key:
            return None, "Clé ANTHROPIC_API_KEY non configurée dans secrets.toml"

        # Convertir PDF en base64
        pdf_b64 = base64.standard_b64encode(raw_bytes).decode("utf-8")

        prompt = """Ce PDF est un relevé bancaire. Extrais TOUTES les lignes du tableau de transactions.
Pour chaque ligne, retourne un objet JSON avec ces clés (si disponibles) :
- date : date de l'opération (format DD/MM/YYYY)
- libelle : description/libellé de l'opération  
- piece : numéro de pièce/référence
- debit : montant débit (nombre positif, 0 si absent)
- credit : montant crédit (nombre positif, 0 si absent)

Ignore les lignes de solde, totaux, en-têtes.
Retourne UNIQUEMENT un tableau JSON valide, rien d'autre. Exemple :
[{"date":"01/01/2026","libelle":"VIREMENT SALAIRE","piece":"VIR001","debit":0,"credit":500000},
 {"date":"02/01/2026","libelle":"CHEQUE 00123","piece":"00123","debit":150000,"credit":0}]"""

        payload = json.dumps({
            "model": "claude-sonnet-4-6",
            "max_tokens": 4000,
            "messages": [{
                "role": "user",
                "content": [
                    {
                        "type": "document",
                        "source": {
                            "type": "base64",
                            "media_type": "application/pdf",
                            "data": pdf_b64
                        }
                    },
                    {"type": "text", "text": prompt}
                ]
            }]
        }).encode("utf-8")

        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        req = urllib.request.Request(
            "https://api.anthropic.com/v1/messages",
            data=payload,
            headers={
                "x-api-key": api_key,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
                "anthropic-beta": "pdfs-2024-09-25"
            },
            method="POST"
        )

        with urllib.request.urlopen(req, context=ctx, timeout=60) as resp:
            result = json.loads(resp.read().decode("utf-8"))

        text_response = result["content"][0]["text"].strip()
        # Nettoyer si besoin
        if text_response.startswith("```"):
            text_response = text_response.split("```")[1]
            if text_response.startswith("json"):
                text_response = text_response[4:]
        text_response = text_response.strip()

        rows = json.loads(text_response)
        if not rows:
            return None, "Aucune transaction détectée dans le relevé"

        df = pd.DataFrame(rows)
        # S'assurer que les colonnes requises existent
        for col in ["date","libelle","piece","debit","credit"]:
            if col not in df.columns:
                df[col] = "" if col in ["date","libelle","piece"] else 0.0
        df["debit"]  = pd.to_numeric(df["debit"],  errors="coerce").fillna(0.0)
        df["credit"] = pd.to_numeric(df["credit"], errors="coerce").fillna(0.0)

        return df, None

    except Exception as ex:
        return None, f"Erreur OCR Claude : {str(ex)[:300]}"


def load_file(uploaded_file):
    name = uploaded_file.name.lower(); uploaded_file.seek(0)
    if name.endswith('.csv'):
        for enc in ['utf-8','latin-1','cp1252']:
            try:
                uploaded_file.seek(0)
                return pd.read_csv(uploaded_file, encoding=enc, header=None, dtype=str)
            except: pass
    elif name.endswith('.pdf'):
        df, err = _extract_pdf_tables(uploaded_file)
        return df  # None si erreur (gérée dans page_import)
    else:
        return pd.read_excel(uploaded_file, header=None, dtype=str)
    return None

def apply_map(df_raw, col_map):
    rows = []; solde_auto = None
    for _, row in df_raw.iterrows():
        D, C = 0.0, 0.0
        if 'montant' in col_map and col_map['montant'] in df_raw.columns:
            mt = pn(row.get(col_map['montant'],'')); ab = abs(mt)
            if ab == 0: continue
            if 'sens' in col_map and col_map['sens'] in df_raw.columns:
                sv = str(row.get(col_map['sens'],'')).lower().strip()
                is_d = sv.startswith('d') or 'débit' in sv or 'debit' in sv or mt < 0
                D, C = (ab, 0.0) if is_d else (0.0, ab)
            else:
                D, C = (ab, 0.0) if mt < 0 else (0.0, ab)
        else:
            d_col = col_map.get('debit'); c_col = col_map.get('credit')
            # Lire valeurs brutes avec signe pour détecter les annulations (valeurs négatives)
            def _pn_signed(val):
                if val is None or str(val).strip() in ('','nan','None','-','—'): return 0.0
                try: return float(str(val).replace(' ','').replace('\xa0','').replace(',','.'))
                except: return 0.0
            d_raw = _pn_signed(row.get(d_col,0)) if d_col and d_col in df_raw.columns else 0.0
            c_raw = _pn_signed(row.get(c_col,0)) if c_col and c_col in df_raw.columns else 0.0
            # Conserver le signe dans la même colonne — abs() pour avoir le montant positif
            # Un montant négatif en CRÉDIT reste en CRÉDIT (valeur absolue)
            # Un montant négatif en DÉBIT reste en DÉBIT (valeur absolue)
            D = abs(d_raw); C = abs(c_raw)
        if D == 0 and C == 0: continue
        lib = str(row.get(col_map.get('lib','__'),'')).strip() if 'lib' in col_map else ''
        if is_excluded(lib):
            if solde_auto is None:
                if C > 0 and D == 0: solde_auto = ('credit', C)
                elif D > 0 and C == 0: solde_auto = ('debit', D)
            continue
        dt = ''
        if 'date' in col_map and col_map['date'] in df_raw.columns:
            s = str(row.get(col_map['date'],'')).strip()
            if s and s not in ('','nan','None'):
                try:
                    d = pd.to_datetime(s, dayfirst=False if re.match(r'^\d{4}[-/]',s) else True, errors='coerce')
                    dt = d.strftime('%d/%m/%Y') if d is not pd.NaT and not pd.isna(d) else s[:10]
                except: dt = s[:10]
        piece = str(row.get(col_map.get('piece','__'),'')).strip() if 'piece' in col_map else ''
        rows.append({'date':dt,'lib':lib,'piece':piece,'debit':D,'credit':C,'matched':False,'match_id':None})
    df = pd.DataFrame(rows) if rows else pd.DataFrame(columns=['date','lib','piece','debit','credit','matched','match_id'])
    return df, solde_auto

# ───────────────────────────────────────────────────────────────────────
# RAPPROCHEMENT
# ───────────────────────────────────────────────────────────────────────
def _extract_numbers(s):
    nums = re.findall(r'\d+', s)
    return {str(int(n)) for n in nums if len(n) >= 2 and int(n) > 0}

def lib_similarity(a, b):
    def normalize(s):
        s = s.lower(); s = re.sub(r'[^a-z0-9]',' ',s)
        return re.sub(r'\s+',' ',s).strip()
    STOP = {'de','du','le','la','les','au','aux','en','et','un','une','des','par','pour','sur',
            'ref','rgl','rglt','nr','n','sa','sarl','suarl','retrait','cheque','compensation','reglement','faveur'}
    def tokens(s): return {t for t in normalize(s).split() if t not in STOP}
    ta, tb = tokens(a), tokens(b)
    if not ta or not tb: return 0.0
    common = ta & tb; long_common = [w for w in common if len(w) > 3]
    base = len(common) / max(len(ta), len(tb))
    na, nb = _extract_numbers(a), _extract_numbers(b)
    if na and nb:
        if na & nb: return min(1.0, base + len(long_common)*0.1 + 0.5)
        else: return 0.0
    return min(1.0, base + len(long_common)*0.1)

def _run_match(rb, cp, inversion):
    rb = rb.copy(); cp = cp.copy()
    rb['matched'] = False; rb['match_id'] = None
    cp['matched'] = False; cp['match_id'] = None
    used = set(); cnt = 0; mc = 0
    for i, rbr in rb.iterrows():
        rD, rC = rbr['debit'], rbr['credit']
        rLib = str(rbr.get('lib',''))
        rb_is_report = _is_report(rbr.to_dict())
        candidates = []
        for j, cpr in cp.iterrows():
            if cpr['matched'] or j in used: continue
            cD, cC = cpr['debit'], cpr['credit']
            ok = False
            if inversion:
                if rD > 0 and abs(rD-cC) <= 1: ok = True
                elif rC > 0 and abs(rC-cD) <= 1: ok = True
            else:
                if rD > 0 and abs(rD-cD) <= 1: ok = True
                elif rC > 0 and abs(rC-cC) <= 1: ok = True
            if not ok: continue
            cp_is_report = _is_report(cpr.to_dict())
            score = 0.5 if (rb_is_report or cp_is_report) else lib_similarity(rLib, str(cpr.get('lib','')))
            candidates.append((j, score))
        if not candidates: continue
        if len(candidates) == 1:
            chosen_j = candidates[0][0]
        else:
            pos = [(j,s) for j,s in candidates if s > 0]
            chosen_j = sorted(pos, key=lambda x:x[1], reverse=True)[0][0] if pos else candidates[0][0]
        mc += 1; cnt += 1; mid = f"M{mc}"
        rb.at[i,'matched'] = True; rb.at[i,'match_id'] = mid
        cp.at[chosen_j,'matched'] = True; cp.at[chosen_j,'match_id'] = mid
        used.add(chosen_j)
    return rb, cp, cnt

def auto_match(rb_df, cp_df, inversion=True):
    st.session_state.mc = getattr(st.session_state, 'mc', 0)

    # ── Séparer carries et courants côté RB ──────────────────────────
    has_carried_rb = 'carried' in rb_df.columns
    if has_carried_rb:
        mask_c     = rb_df['carried'].fillna(False).astype(bool)
        rb_courant = rb_df[~mask_c].copy().reset_index(drop=True)
        rb_carry   = rb_df[ mask_c].copy().reset_index(drop=True)
    else:
        rb_courant = rb_df.copy().reset_index(drop=True)
        rb_carry   = pd.DataFrame(columns=rb_df.columns if len(rb_df) else
                                  ['date','lib','piece','debit','credit','matched','match_id'])

    # ── Séparer carries et courants côté GL ──────────────────────────
    has_carried_cp = 'carried' in cp_df.columns
    if has_carried_cp:
        mask_cp_c   = cp_df['carried'].fillna(False).astype(bool)
        cp_courant  = cp_df[~mask_cp_c].copy().reset_index(drop=True)
        cp_carry    = cp_df[ mask_cp_c].copy().reset_index(drop=True)
    else:
        cp_courant  = cp_df.copy().reset_index(drop=True)
        cp_carry    = pd.DataFrame(columns=cp_df.columns if len(cp_df) else
                                   ['date','lib','piece','debit','credit','matched','match_id'])

    # ── PASSE 0 : carry GL ↔ cp_courant (même sens/montant) ──────────
    # Un suspens GL CRÉDIT reporté (ex: chèque émis en fév) peut avoir sa
    # contrepartie CRÉDIT dans le GL mars (même écriture re-comptabilisée).
    # On les régularise : les deux disparaissent complètement.
    p0_cnt = 0; cp_reserved_p0 = set(); cp_carry_reserved = set()
    if len(cp_carry) > 0:
        cp_carry['matched']  = False
        cp_carry['match_id'] = None
        for ci in cp_carry.index:
            carry_d = cp_carry.at[ci, 'debit']
            carry_c = cp_carry.at[ci, 'credit']
            if carry_d == 0 and carry_c == 0: continue
            for ri in cp_courant.index:
                if ri in cp_reserved_p0: continue
                cur_d = cp_courant.at[ri, 'debit']
                cur_c = cp_courant.at[ri, 'credit']
                # Même sens : crédit↔crédit OU débit↔débit
                if (carry_c > 0 and abs(carry_c - cur_c) <= 1) or                    (carry_d > 0 and abs(carry_d - cur_d) <= 1):
                    mid = f"CP0_{ci}_{ri}"
                    cp_carry.at[ci, 'matched']    = True
                    cp_carry.at[ci, 'match_id']   = mid
                    cp_courant.at[ri, 'matched']  = True
                    cp_courant.at[ri, 'match_id'] = mid
                    cp_reserved_p0.add(ri)
                    cp_carry_reserved.add(ci)
                    p0_cnt += 1; break

    # Carries GL non régularisés → restent comme suspens GL reportés
    cp_carry_non_reg = cp_carry[~cp_carry['matched'].fillna(False)].copy() if len(cp_carry) > 0 else pd.DataFrame(columns=cp_df.columns)
    # cp_df pour les passes suivantes = courants libres + carries non régularisés
    cp_df_for_match = pd.concat(
        [cp_courant[~cp_courant['matched'].fillna(False)], cp_carry_non_reg],
        ignore_index=True
    )

    # ── PASSE 1 : carryover_rb ↔ rb_courant (même sens, même montant) ──
    rb_carry_out = rb_carry.copy()
    p1_cnt = 0; rb_reserved = set()

    if len(rb_carry_out) > 0:
        rb_carry_out['matched']  = False
        rb_carry_out['match_id'] = None
        mc1 = st.session_state.mc
        for ci in rb_carry_out.index:
            carry_d = rb_carry_out.at[ci,'debit']
            carry_c = rb_carry_out.at[ci,'credit']
            if carry_d == 0 and carry_c == 0: continue
            for ri in rb_courant.index:
                if ri in rb_reserved: continue
                cur_d = rb_courant.at[ri,'debit']
                cur_c = rb_courant.at[ri,'credit']
                if (carry_d > 0 and abs(carry_d - cur_d) <= 1) or                    (carry_c > 0 and abs(carry_c - cur_c) <= 1):
                    mc1 += 1; mid = f"C{mc1}"
                    rb_carry_out.at[ci,'matched']  = True
                    rb_carry_out.at[ci,'match_id'] = mid
                    rb_courant.at[ri,'matched']    = True
                    rb_courant.at[ri,'match_id']   = mid
                    rb_reserved.add(ri)
                    p1_cnt += 1; break

    # ── PASSE 1b : carry_rb non régularisés ↔ cp_df (inversion D/C) ──
    cp_reserved = set()
    if len(rb_carry_out) > 0:
        mc1b = st.session_state.mc
        cp_tmp = cp_df_for_match.copy().reset_index(drop=True)
        for ci in rb_carry_out.index:
            if rb_carry_out.at[ci, 'matched']: continue
            carry_d = rb_carry_out.at[ci, 'debit']
            carry_c = rb_carry_out.at[ci, 'credit']
            if carry_d == 0 and carry_c == 0: continue
            for cj in cp_tmp.index:
                if cp_tmp.at[cj, 'matched'] or cj in cp_reserved: continue
                cp_d = cp_tmp.at[cj, 'debit']
                cp_c = cp_tmp.at[cj, 'credit']
                matched_inv = (carry_d > 0 and abs(carry_d - cp_c) <= 1) or                               (carry_c > 0 and abs(carry_c - cp_d) <= 1)
                if matched_inv:
                    mc1b += 1; mid = f"P{mc1b}"
                    rb_carry_out.at[ci, 'matched']  = True
                    rb_carry_out.at[ci, 'match_id'] = mid
                    cp_tmp.at[cj, 'matched']        = True
                    cp_tmp.at[cj, 'match_id']       = mid
                    cp_reserved.add(cj)
                    p1_cnt += 1; break
        cp_df_for_match = cp_tmp

    # ── PASSE 2 : rb_courant libres ↔ cp_df_for_match (3 stratégies) ──
    rb_libre = rb_courant[~rb_courant['matched']].copy().reset_index(drop=True)
    rb_pris  = rb_courant[ rb_courant['matched']].copy().reset_index(drop=True)

    rb2a, cp2a, cnt2a = _run_match(rb_libre, cp_df_for_match, inversion)
    rb2b, cp2b, cnt2b = _run_match(rb_libre, cp_df_for_match, not inversion)
    rb2c_a, cp2c_a, _ = _run_match(rb_libre, cp_df_for_match, True)
    rb2c_u = rb2c_a[~rb2c_a['matched']].copy(); cp2c_u = cp2c_a[~cp2c_a['matched']].copy()
    if len(rb2c_u) > 0 and len(cp2c_u) > 0:
        rb2c_b, cp2c_b, _ = _run_match(rb2c_u, cp2c_u, False)
        rb2c_a.update(rb2c_b); cp2c_a.update(cp2c_b)
    cnt2c = int(rb2c_a['matched'].sum())
    p2_cnt, rb_p2, cp_p2 = max([(cnt2a,rb2a,cp2a),(cnt2b,rb2b,cp2b),(cnt2c,rb2c_a,cp2c_a)], key=lambda x:x[0])

    # ── Fusionner ─────────────────────────────────────────────────────
    # rb_p2            : lignes courantes libres (matchées GL ou non → suspens ou rapprochées)
    # rb_pris          : lignes courant consommées par Passe 1 → EXCLUES COMPLÈTEMENT
    #                    (elles sont régularisées avec un carry, elles n'ont rien à faire
    #                     dans rb_df ni dans le tableau ERB)
    # rb_carry_non_reg : carries non régularisés → suspens rouges dans le tableau ERB
    rb_carry_non_reg = rb_carry_out[~rb_carry_out['matched'].fillna(False).astype(bool)].copy()
    parts = [rb_p2]   # rb_pris intentionnellement exclu
    if len(rb_carry_non_reg) > 0:
        parts.append(rb_carry_non_reg)
    rb_out = pd.concat(parts, ignore_index=True)
    # cp_out = résultat passe 2 (déjà filtré des carries régularisés en passe 0)
    cp_out = cp_p2
    total  = p1_cnt + p2_cnt

    mc = st.session_state.mc
    all_ids = set(rb_out[rb_out['matched']]['match_id'].tolist() +
                  cp_out[cp_out['matched']]['match_id'].tolist())
    mapping = {old: f"A{mc+i+1}" for i, old in enumerate(sorted(all_ids))}
    rb_out['match_id'] = rb_out['match_id'].map(lambda x: mapping.get(x, x))
    cp_out['match_id'] = cp_out['match_id'].map(lambda x: mapping.get(x, x))

    st.session_state.mc += total
    return rb_out.reset_index(drop=True), cp_out.reset_index(drop=True), total

# ───────────────────────────────────────────────────────────────────────
# GESTION DES REPORTS INTER-MOIS
# ───────────────────────────────────────────────────────────────────────
def _is_report(r):
    if r is None: return False
    cf = r.get('carry_from') if isinstance(r, dict) else None
    return bool(cf and str(cf).strip() not in ('','nan','None','False'))

def get_rb_with_carryover():
    rb = st.session_state.rb_df
    carryover = st.session_state.carryover_rb
    if not carryover: return rb
    df_carry = pd.DataFrame(carryover); df_carry['carried'] = True
    if rb is not None and len(rb):
        if 'carried' not in rb.columns: rb = rb.copy(); rb['carried'] = False
        return pd.concat([rb, df_carry], ignore_index=True)
    return df_carry

def get_cp_with_carryover():
    cp = st.session_state.cp_df
    carryover = st.session_state.carryover_cp
    if not carryover: return cp
    df_carry = pd.DataFrame(carryover); df_carry['carried'] = True
    if cp is not None and len(cp):
        if 'carried' not in cp.columns: cp = cp.copy(); cp['carried'] = False
        return pd.concat([cp, df_carry], ignore_index=True)
    return df_carry

def extraire_regularisees(rb_matched, cp_matched):
    regularisees = []
    if 'carry_from' in rb_matched.columns:
        for _, r in rb_matched.iterrows():
            if _is_report(r.to_dict()) and r.get('matched'):
                mid = r.get('match_id')
                partner = cp_matched[cp_matched['match_id'] == mid] if 'match_id' in cp_matched.columns else pd.DataFrame()
                regularisees.append({
                    'type': 'RB régularisé', 'from': r.get('carry_from',''),
                    'date': r.get('date',''), 'lib': r.get('lib',''),
                    'debit': r.get('debit',0), 'credit': r.get('credit',0),
                    'partner_lib': partner.iloc[0]['lib'] if len(partner) else '(même relevé)',
                    'partner_date': partner.iloc[0]['date'] if len(partner) else '',
                })
    if 'carry_from' in cp_matched.columns:
        for _, r in cp_matched.iterrows():
            if _is_report(r.to_dict()) and r.get('matched'):
                mid = r.get('match_id')
                partner = rb_matched[rb_matched['match_id'] == mid]
                regularisees.append({
                    'type': 'GL régularisé', 'from': r.get('carry_from',''),
                    'date': r.get('date',''), 'lib': r.get('lib',''),
                    'debit': r.get('debit',0), 'credit': r.get('credit',0),
                    'partner_lib': partner.iloc[0]['lib'] if len(partner) else '',
                    'partner_date': partner.iloc[0]['date'] if len(partner) else '',
                })
    return regularisees

def cloturer_mois(periode, rb_df, cp_df, erb_html_str=''):
    susp_rb = rb_df[~rb_df['matched']].copy() if rb_df is not None else pd.DataFrame()
    susp_cp = cp_df[~cp_df['matched']].copy() if cp_df is not None else pd.DataFrame()
    def prep_carry(df):
        if df.empty: return []
        df = df.copy(); df['matched'] = False; df['match_id'] = None; df['carried'] = True
        if 'carry_from' not in df.columns:
            df['carry_from'] = periode
        else:
            df['carry_from'] = df['carry_from'].apply(
                lambda x: x if (x and str(x).strip() not in ('','nan','None','False')) else periode)
        return df.to_dict('records')
    st.session_state.carryover_rb = prep_carry(susp_rb)
    st.session_state.carryover_cp = prep_carry(susp_cp)
    st.session_state.regularisees = []
    st.session_state.historique.append({
        'periode': periode, 'n_rb': len(susp_rb), 'n_cp': len(susp_cp),
        'susp_rb': susp_rb.to_dict('records') if not susp_rb.empty else [],
        'susp_cp': susp_cp.to_dict('records') if not susp_cp.empty else [],
        'erb_html': erb_html_str,
    })
    # Réinitialiser pour le mois suivant
    st.session_state.rb_df = None; st.session_state.cp_df = None
    st.session_state.rb_solde_auto = None; st.session_state.cp_solde_auto = None
    st.session_state.mc = 0
    st.session_state.s_rb = 0.0; st.session_state.s_cp = 0.0
    # Supprimer les clés widget (ne pas assigner directement = erreur Streamlit)
    for _k in ['inf_pe', 'inf_cl', 'inf_bq', 'inf_co',
               '_inp_s_rb', '_inp_s_cp']:
        if _k in st.session_state: del st.session_state[_k]

# ───────────────────────────────────────────────────────────────────────
# CALCUL ERB  (logique v5.5 validée)
# ───────────────────────────────────────────────────────────────────────
def calc_erb(rb_df, cp_df, s_rb, s_cp):
    """
    Convention ERB Ecobank — logique validée sur ERB de reference :

    COTE RELEVE col D = suspens RB debits (reports rouges + courants)
                      + suspens GL credits COURANTS inversés (cheques emis sans compensation)
    COTE RELEVE col E = solde RB crediteur
                      + suspens GL debits courants inversés (GL debit sans RB credit)
    SR_RB = col E - col D

    COTE JOURNAL col I = solde GL debiteur
    COTE JOURNAL col J = suspens GL credits REPORTÉS (cheques rejetes, valeur abs)
                       + suspens RB credits courants
    SR_CP = col I - col J  =>  solde GL + |cheques rejetes|

    Equilibre : SR_RB = SR_CP  =>  chk = SR_RB - SR_CP ~ 0
    Distinction courant/reporté : _is_report(r) via carry_from non vide
    """
    susp_rb = rb_df[~rb_df['matched']].reset_index(drop=True) if rb_df is not None and len(rb_df) else pd.DataFrame()
    susp_cp = cp_df[~cp_df['matched']].reset_index(drop=True) if cp_df is not None and len(cp_df) else pd.DataFrame()

    susp_rb_debit  = susp_rb[(susp_rb['debit']>0)&(susp_rb['credit']==0)].reset_index(drop=True) if not susp_rb.empty else pd.DataFrame()
    susp_rb_credit = susp_rb[(susp_rb['credit']>0)&(susp_rb['debit']==0)].reset_index(drop=True) if not susp_rb.empty else pd.DataFrame()
    susp_cp_debit  = susp_cp[(susp_cp['debit']>0)&(susp_cp['credit']==0)].reset_index(drop=True) if not susp_cp.empty else pd.DataFrame()
    susp_cp_credit = susp_cp[(susp_cp['credit']>0)&(susp_cp['debit']==0)].reset_index(drop=True) if not susp_cp.empty else pd.DataFrame()

    # Distinguer GL credits courants (cheques emis) vs reportes (cheques rejetes)
    if not susp_cp_credit.empty:
        mask_rep = susp_cp_credit.apply(lambda r: _is_report(r.to_dict()), axis=1)
        susp_cp_credit_courant = susp_cp_credit[~mask_rep].reset_index(drop=True)
        susp_cp_credit_report  = susp_cp_credit[ mask_rep].reset_index(drop=True)
    else:
        susp_cp_credit_courant = pd.DataFrame()
        susp_cp_credit_report  = pd.DataFrame()

    # ═══════════════════════════════════════════════════════════════════
    # CONVENTION ERB ECOBANK — FORMULES MATHÉMATIQUEMENT EXACTES
    # ───────────────────────────────────────────────────────────────────
    # Identité Ecobank : S_RB = S_CP + ΣRB_D + ΣGL_C
    #
    # CÔTÉ RELEVÉ :
    #   col D = ΣRB_D  (suspens DÉBIT relevé : chèques tirés non encore compensés)
    #   col E = S_RB   (solde final relevé)
    #   SR_RB = col E - col D = S_RB - ΣRB_D
    #
    # CÔTÉ JOURNAL :
    #   col I = S_CP   (solde final GL)
    #   col J = ΣGL_C  (suspens CRÉDIT GL : chèques rejetés, affiché négatif)
    #   SR_CP = col I + col J = S_CP + ΣGL_C
    #
    # SR_RB = SR_CP toujours si S_RB = S_CP + ΣRB_D + ΣGL_C
    #
    # ΣRB_D = suspens relevé DÉBIT  (all : courants + reportés)
    # ΣRB_C = suspens relevé CRÉDIT (versements relevé sans GL — non inclus dans SR)
    # ΣGL_C = suspens GL CRÉDIT     (chèques rejetés, courants + reportés)
    # ΣGL_D = suspens GL DÉBIT      (encaissements GL sans relevé — non inclus dans SR)
    # ═══════════════════════════════════════════════════════════════════

    # Calcul des sommes de suspens
    sum_susp_rb_d = susp_rb_debit['debit'].sum()            if not susp_rb_debit.empty          else 0.0
    sum_susp_rb_c = susp_rb_credit['credit'].sum()          if not susp_rb_credit.empty         else 0.0
    sum_susp_gl_d = susp_cp_debit['debit'].sum()            if not susp_cp_debit.empty          else 0.0
    sum_susp_gl_c = (susp_cp_credit_courant['credit'].sum() if not susp_cp_credit_courant.empty else 0.0) +                     (susp_cp_credit_report['credit'].sum()  if not susp_cp_credit_report.empty  else 0.0)

    # ═══════════════════════════════════════════════════════════════════
    # CONVENTION ERB ECOBANK — FORMULES EXACTES (validées sur ERB manuel)
    # ───────────────────────────────────────────────────────────────────
    # Identité : S_RB = S_CP + ΣRB_D + ΣGL_C_courant + ΣGL_C_reporté
    #
    # Col D (côté relevé) = ΣRB_D + ΣGL_C_courant
    #   (suspens DÉBIT relevé + chèques émis GL non compensés au relevé)
    # Col E (côté relevé) = S_RB
    #   SR_RB = S_RB - ΣRB_D - ΣGL_C_courant
    #
    # Col I (côté journal) = S_CP
    # Col J (côté journal) = ΣGL_C_reporté  (chèques rejetés reportés, affiché négatif)
    #   SR_CP = S_CP + ΣGL_C_reporté
    #
    # SR_RB = SR_CP toujours si S_RB = S_CP + ΣRB_D + ΣGL_C (toutes ΣGL_C)
    # ═══════════════════════════════════════════════════════════════════

    sum_susp_gl_c_courant = susp_cp_credit_courant['credit'].sum() if not susp_cp_credit_courant.empty else 0.0
    sum_susp_gl_c_report  = susp_cp_credit_report['credit'].sum()  if not susp_cp_credit_report.empty  else 0.0

    # Soldes rapprochés — convention Ecobank exacte
    sr_rb = s_rb - sum_susp_rb_d - sum_susp_gl_c_courant
    sr_cp = s_cp + sum_susp_gl_c_report

    # Variables de présentation tableau ERB
    rb_sol_d = abs(s_rb) if s_rb < 0 else 0.0
    rb_sol_c = s_rb      if s_rb >= 0 else 0.0
    cp_sol_d = s_cp      if s_cp >= 0 else 0.0
    cp_sol_c = abs(s_cp) if s_cp <  0 else 0.0
    cp_susp_d_rb     = sum_susp_rb_c
    cp_susp_c_report = sum_susp_gl_c_report

    # Totaux affichage tableau
    rb_susp_d_rb         = sum_susp_rb_d
    rb_susp_d_gl_courant = sum_susp_gl_c_courant  # GL crédits courants → col D relevé
    rb_susp_c_gl_debit   = sum_susp_gl_d
    tot_d_rb = rb_sol_d + rb_susp_d_rb + rb_susp_d_gl_courant
    tot_c_rb = rb_sol_c
    tot_d_cp = cp_sol_d
    tot_c_cp = cp_sol_c + sum_susp_gl_c_report   # col J = ΣGL_C reportés seulement

    chk = sr_rb - sr_cp; ok = abs(chk) < 0.5
    return dict(
        susp_rb=susp_rb, susp_cp=susp_cp,
        susp_rb_debit=susp_rb_debit, susp_rb_credit=susp_rb_credit,
        susp_cp_debit=susp_cp_debit, susp_cp_credit=susp_cp_credit,
        susp_cp_credit_courant=susp_cp_credit_courant,
        susp_cp_credit_report=susp_cp_credit_report,
        rb_sol_d=rb_sol_d, rb_sol_c=rb_sol_c, cp_sol_d=cp_sol_d, cp_sol_c=cp_sol_c,
        tot_d_rb=tot_d_rb, tot_c_rb=tot_c_rb, sr_rb=sr_rb,
        tot_d_cp=tot_d_cp, tot_c_cp=tot_c_cp, sr_cp=sr_cp, chk=chk, ok=ok,
        cp_susp_d_rb=cp_susp_d_rb,
    )

def build_erb_html(e, info):
    """
    COTE RELEVE :
      rows_left = GL debits (inverses, col E) + RB debits + GL credits COURANTS (col D)
    COTE JOURNAL :
      rows_right = GL credits REPORTES (cheques rejetes, col J) + RB credits
    """
    susp_cp_debit          = e.get('susp_cp_debit', pd.DataFrame())
    susp_cp_credit_courant = e.get('susp_cp_credit_courant', pd.DataFrame())
    susp_cp_credit_report  = e.get('susp_cp_credit_report', pd.DataFrame())
    susp_rb_debit  = e['susp_rb_debit']
    susp_rb_credit = e['susp_rb_credit']

    # COTE RELEVE col D : RB debits (reports rouges en premier) + GL credits courants inverses
    # COTE RELEVE col E : GL debits inverses en credit
    rows_left = []
    # GL debits → col E (credit releve) — affiches en bleu clair, valeur en CREDIT
    for _, r in susp_cp_debit.iterrows():
        rows_left.append({'row': r, 'side': 'gl_debit', 'is_rep': _is_report(r.to_dict())})
    # RB debits → col D (debit releve)
    for _, r in susp_rb_debit.iterrows():
        rows_left.append({'row': r, 'side': 'rb_debit', 'is_rep': _is_report(r.to_dict())})
    # GL credits courants → col D (debit releve, inverses)
    for _, r in susp_cp_credit_courant.iterrows():
        rows_left.append({'row': r, 'side': 'gl_credit_courant', 'is_rep': False})

    # COTE JOURNAL :
    # col I (DÉBIT)  : solde GL + suspens RB crédits (versements crédit relevé → débit journal)
    # col J (CRÉDIT) : GL crédits REPORTÉS seulement (chèques rejetés du mois précédent)
    rows_right_debit  = []  # col I — DÉBIT journal
    rows_right_credit = []  # col J — CRÉDIT journal
    # Suspens RB crédits → DÉBIT journal (col I)
    for _, r in susp_rb_credit.iterrows():
        rows_right_debit.append({'row': r, 'side': 'rb_credit_as_debit', 'is_rep': _is_report(r.to_dict())})
    # GL crédits reportés → CRÉDIT journal (col J)
    for _, r in susp_cp_credit_report.iterrows():
        rows_right_credit.append({'row': r, 'side': 'gl_credit_report', 'is_rep': True})
    # Fusionner pour l'affichage ligne par ligne
    rows_right = []
    max_right = max(len(rows_right_debit), len(rows_right_credit))
    for i in range(max_right):
        d_item = rows_right_debit[i]  if i < len(rows_right_debit)  else None
        c_item = rows_right_credit[i] if i < len(rows_right_credit) else None
        rows_right.append({'debit_item': d_item, 'credit_item': c_item})

    n_rows = max(len(rows_left), len(rows_right), 8)
    periode = info.get('pe','')

    S = dict(
        TH  = 'background:#4472C4;color:#fff;font-weight:700;padding:6px 8px;border:1px solid #2E5A9C;text-align:center;font-size:11px;white-space:nowrap',
        THL = 'background:#4472C4;color:#fff;font-weight:700;padding:6px 8px;border:1px solid #2E5A9C;text-align:left;font-size:11px',
        RB  = 'padding:5px 8px;border:.5px solid #7BA7D0;font-size:11px;background:#C5D9F1;color:#0A2540',
        RBR = 'padding:5px 8px;border:.5px solid #7BA7D0;font-size:11px;background:#C5D9F1;text-align:right;color:#0A2540',
        ER  = 'padding:5px 8px;border:.5px solid #B8D0E8;font-size:11px;background:#E8F2FA;color:#0A2540',
        CP  = 'padding:5px 8px;border:.5px solid #C49070;font-size:11px;background:#F2CBAB;color:#3A1800',
        CPR = 'padding:5px 8px;border:.5px solid #C49070;font-size:11px;background:#F2CBAB;text-align:right;color:#3A1800',
        EC  = 'padding:5px 8px;border:.5px solid #E0C4A0;font-size:11px;background:#FAF0E6;color:#3A1800',
        FT  = 'padding:6px 8px;border:1px solid #4A8ABB;font-size:11px;font-weight:700;background:#6A9FCC;color:#fff;text-align:right',
        FTL = 'padding:6px 8px;border:1px solid #4A8ABB;font-size:11px;font-weight:700;background:#6A9FCC;color:#fff;text-align:center',
        SR  = 'padding:6px 8px;border:1px solid #0090C8;font-size:12px;font-weight:700;background:#0099D8;color:#fff;text-align:right',
        SRL = 'padding:6px 8px;border:1px solid #0090C8;font-size:12px;font-weight:700;background:#0099D8;color:#fff;text-align:left',
        SOK = 'padding:7px 8px;border:1px solid #4A8A20;font-size:12px;font-weight:700;background:#5A9E28;color:#fff;text-align:center',
        SER = 'padding:7px 8px;border:1px solid #C00;font-size:12px;font-weight:700;background:#D32F2F;color:#fff;text-align:center',
        LB  = 'padding:5px 8px;border:.5px solid #bbb;font-size:11px;font-weight:700;background:#E8E8E8;color:#1A1A1A',
        VL  = 'padding:5px 8px;border:.5px solid #bbb;font-size:11px;background:#FFFFFF;color:#1A1A1A',
        RR  = 'padding:5px 8px;border:1px solid #C00;font-size:11px;background:#FFAAAA;font-weight:700;color:#7A0000',
        RRR = 'padding:5px 8px;border:1px solid #C00;font-size:11px;background:#FFAAAA;font-weight:700;text-align:right;color:#7A0000',
    )

    h = '<table style="width:100%;border-collapse:collapse;font-size:11px;min-width:700px;color:#1a1a1a;background:#fff">'
    h += f'<tr><td colspan="10" style="text-align:center;font-weight:700;font-size:13px;padding:9px;background:#4472C4;color:#fff;border:1px solid #2E5A9C;letter-spacing:.06em">RAPPROCHEMENT BANCAIRE</td></tr>'
    h += f'<tr><td colspan="2" style="{S["LB"]}">CLIENT</td><td colspan="3" style="{S["VL"]}">{info.get("cl","")}</td><td colspan="2" style="{S["LB"]}">BANQUE</td><td colspan="3" style="{S["VL"]}">{info.get("bq","")}</td></tr>'
    h += f'<tr><td colspan="2" style="{S["LB"]}">N° COMPTE</td><td colspan="3" style="{S["VL"]}">{info.get("co","")}</td><td colspan="2" style="{S["LB"]}">PÉRIODE</td><td colspan="3" style="{S["VL"]}">{periode}</td></tr>'
    h += f'<tr><td colspan="5" style="{S["TH"]}">RELEVE BANCAIRE</td><td colspan="5" style="{S["TH"]}">JOURNAL BANQUE</td></tr>'
    h += (f'<tr><td style="{S["TH"]}">DATE</td><td style="{S["THL"]}">LIBELLE</td>'
          f'<td style="{S["TH"]}">N° PIECE</td><td style="{S["TH"]}">DEBIT</td><td style="{S["TH"]}">CREDIT</td>'
          f'<td style="{S["TH"]}">DATE</td><td style="{S["THL"]}">LIBELLE</td>'
          f'<td style="{S["TH"]}">N° PIECE</td><td style="{S["TH"]}">DEBIT</td><td style="{S["TH"]}">CREDIT</td></tr>')

    rb_sd = fmt_fr(e['rb_sol_d']) if e['rb_sol_d'] > 0 else ''
    rb_sc = fmt_fr(e['rb_sol_c']) if e['rb_sol_c'] > 0 else ''
    cp_sd = fmt_fr(e['cp_sol_d']) if e['cp_sol_d'] > 0 else ''
    cp_sc = fmt_fr(e['cp_sol_c']) if e['cp_sol_c'] > 0 else ''
    h += f'<tr><td style="{S["RB"]}"></td><td colspan="2" style="{S["RB"]}">SOLDE AVANT RAPPROCHEMENT AU {periode}</td>'
    h += f'<td style="{S["RBR"]}"><b>{rb_sd}</b></td><td style="{S["RBR"]}"><b>{rb_sc}</b></td>'
    h += f'<td style="{S["CP"]}"></td><td colspan="2" style="{S["CP"]}">SOLDE AVANT RAPPROCHEMENT AU {periode}</td>'
    h += f'<td style="{S["CPR"]}"><b>{cp_sd}</b></td><td style="{S["CPR"]}"><b>{cp_sc}</b></td></tr>'

    for i in range(n_rows):
        # ── COTE RELEVE ──────────────────────────────────────────────
        if i < len(rows_left):
            item = rows_left[i]; r = item['row']; side = item['side']; is_rep = item['is_rep']
            lb = r['lib']
            if is_rep: lb = f'[↩ {r.get("carry_from","")}] {lb}'
            s_tx = S['RR'] if is_rep else S['RB']
            s_num = S['RRR'] if is_rep else S['RBR']
            if side == 'gl_debit':
                # GL debit → col E (credit releve, inversi)
                d_val = ''; c_val = fmt_fr(r['debit']) if r['debit'] > 0 else ''
            elif side == 'rb_debit':
                # RB debit → col D
                d_val = fmt_fr(r['debit']) if r['debit'] > 0 else ''; c_val = ''
            else:  # gl_credit_courant
                # GL credit courant → col D (debit releve, inverse)
                d_val = fmt_fr(r['credit']) if r['credit'] > 0 else ''; c_val = ''
            h += (f'<tr><td style="{s_tx}">{r["date"]}</td><td style="{s_tx}">{lb}</td>'
                  f'<td style="{s_tx}">{r.get("piece","")}</td>'
                  f'<td style="{s_num}">{d_val}</td><td style="{s_num}">{c_val}</td>')
        else:
            h += (f'<tr><td style="{S["ER"]}"></td><td style="{S["ER"]}"></td>'
                  f'<td style="{S["ER"]}"></td><td style="{S["ER"]}"></td><td style="{S["ER"]}"></td>')

        # ── COTE JOURNAL ─────────────────────────────────────────────
        if i < len(rows_right):
            pair = rows_right[i]
            d_item = pair.get('debit_item')   # col I — DÉBIT
            c_item = pair.get('credit_item')  # col J — CRÉDIT
            if d_item:
                r2 = d_item['row']; is_rep2 = d_item['is_rep']
                lb2 = r2['lib']
                if is_rep2: lb2 = f'[↩ {r2.get("carry_from","")}] {lb2}'
                s_tx2 = S['RR'] if is_rep2 else S['CP']
                s_num2 = S['RRR'] if is_rep2 else S['CPR']
                # RB crédit → DÉBIT journal (col I) — montant en débit
                d2 = fmt_fr(r2['credit']) if r2['credit'] > 0 else ''
                c2 = fmt_fr(c_item['row']['credit']) if c_item and c_item['row']['credit'] > 0 else ''
                lb_c = ''
                if c_item:
                    lb_c_r = c_item['row']['lib']
                    if c_item['is_rep']: lb_c_r = f'[↩ {c_item["row"].get("carry_from","")}] {lb_c_r}'
                h += (f'<td style="{s_tx2}">{r2["date"]}</td><td style="{s_tx2}">{lb2}</td>'
                      f'<td style="{s_tx2}">{r2.get("piece","")}</td>'
                      f'<td style="{s_num2}">{d2}</td><td style="{s_num2}">{c2}</td></tr>')
            elif c_item:
                r2 = c_item['row']; is_rep2 = c_item['is_rep']
                lb2 = r2['lib']
                if is_rep2: lb2 = f'[↩ {r2.get("carry_from","")}] {lb2}'
                s_tx2 = S['RR'] if is_rep2 else S['CP']
                s_num2 = S['RRR'] if is_rep2 else S['CPR']
                c2 = fmt_fr(r2['credit']) if r2['credit'] > 0 else ''
                h += (f'<td style="{s_tx2}">{r2["date"]}</td><td style="{s_tx2}">{lb2}</td>'
                      f'<td style="{s_tx2}">{r2.get("piece","")}</td>'
                      f'<td style="{s_num2}"></td><td style="{s_num2}">{c2}</td></tr>')
            else:
                h += (f'<td style="{S["EC"]}"></td><td style="{S["EC"]}"></td>'
                      f'<td style="{S["EC"]}"></td><td style="{S["EC"]}"></td><td style="{S["EC"]}"></td></tr>')
        else:
            h += (f'<td style="{S["EC"]}"></td><td style="{S["EC"]}"></td>'
                  f'<td style="{S["EC"]}"></td><td style="{S["EC"]}"></td><td style="{S["EC"]}"></td></tr>')

    h += (f'<tr><td colspan="3" style="{S["FTL"]}">TOTAUX</td>'
          f'<td style="{S["FT"]}">{fmt_fr(e["tot_d_rb"]) or "0,00"}</td>'
          f'<td style="{S["FT"]}">{fmt_fr(e["tot_c_rb"]) or "0,00"}</td>'
          f'<td colspan="3" style="{S["FTL"]}">TOTAUX</td>'
          f'<td style="{S["FT"]}">{fmt_fr(e["tot_d_cp"]) or "0,00"}</td>'
          f'<td style="{S["FT"]}">{fmt_fr(e["tot_c_cp"]) if e["tot_c_cp"] else "-"}</td></tr>')

    h += (f'<tr><td style="{S["SR"]}"></td><td colspan="2" style="{S["SRL"]}">SOLDES RAPPROCHES</td>'
          f'<td style="{S["SR"]}"><b>{fmt_fr(e["sr_rb"])}</b></td><td style="{S["SR"]}"></td>'
          f'<td style="{S["SR"]}"></td><td colspan="2" style="{S["SRL"]}">SOLDES RAPPROCHES</td>'
          f'<td style="{S["SR"]}"></td><td style="{S["SR"]}"><b>{fmt_fr(e["sr_cp"])}</b></td></tr>')

    st_s = S['SOK'] if e['ok'] else S['SER']
    st_t = 'OK' if e['ok'] else 'A RAPPROCHER'
    h += (f'<tr><td colspan="4" style="padding:5px 8px;border:.5px solid #ddd;background:#fff"></td>'
          f'<td colspan="2" style="padding:5px 8px;border:.5px solid #ddd;font-size:11px;font-weight:600;background:#E2EFDA;text-align:center">Statut</td>'
          f'<td colspan="4" style="{st_s}">{st_t}</td></tr>')
    h += '</table>'
    return h

def export_xlsx(e, info):
    susp_cp_debit          = e.get('susp_cp_debit', pd.DataFrame())
    susp_cp_credit_courant = e.get('susp_cp_credit_courant', pd.DataFrame())
    susp_cp_credit_report  = e.get('susp_cp_credit_report', pd.DataFrame())
    susp_rb_debit  = e['susp_rb_debit']
    susp_rb_credit = e['susp_rb_credit']

    rows_left = []
    for _, r in susp_cp_debit.iterrows():
        rows_left.append({'row': r, 'side': 'gl_debit', 'is_rep': _is_report(r.to_dict())})
    for _, r in susp_rb_debit.iterrows():
        rows_left.append({'row': r, 'side': 'rb_debit', 'is_rep': _is_report(r.to_dict())})
    for _, r in susp_cp_credit_courant.iterrows():
        rows_left.append({'row': r, 'side': 'gl_credit_courant', 'is_rep': False})
    # col I (DÉBIT journal)  : suspens RB crédits → débit journal
    # col J (CRÉDIT journal) : GL crédits reportés seulement
    rows_right_debit  = []
    rows_right_credit = []
    for _, r in susp_rb_credit.iterrows():
        rows_right_debit.append({'row': r, 'side': 'rb_credit_as_debit', 'is_rep': _is_report(r.to_dict())})
    for _, r in susp_cp_credit_report.iterrows():
        rows_right_credit.append({'row': r, 'side': 'gl_credit_report', 'is_rep': True})
    max_right = max(len(rows_right_debit), len(rows_right_credit))
    rows_right = []
    for i in range(max_right):
        rows_right.append({
            'debit_item':  rows_right_debit[i]  if i < len(rows_right_debit)  else None,
            'credit_item': rows_right_credit[i] if i < len(rows_right_credit) else None,
        })

    n_rows = max(len(rows_left), len(rows_right), 8)
    periode = info.get('pe','')
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "ERB"
    def fill(h): return PatternFill("solid", fgColor=h)
    def tb():
        s = Side(border_style="thin", color="000000")
        return Border(left=s, right=s, top=s, bottom=s)
    def cell(r, c, v=None, fill_=None, bold=False, color="000000", align="left", border=True):
        cl = ws.cell(row=r, column=c, value=v)
        if fill_: cl.fill = fill_
        cl.font = Font(bold=bold, color=color, size=10)
        cl.alignment = Alignment(horizontal=align, vertical='center', wrap_text=False)
        if border: cl.border = tb()
        return cl
    BH=fill("4472C4"); BL=fill("C5D9F1"); BE=fill("DAE9F7")
    OL=fill("F2CBAB"); OE=fill("FCE4D6"); ST=fill("8EB4E3")
    CY=fill("00B0F0"); GR=fill("92D050"); RD=fill("FF0000")
    WH=fill("FFFFFF"); GY=fill("F0F0F0"); RL=fill("FFCCCC")

    row = 1
    ws.merge_cells(f'A{row}:J{row}')
    cell(row,1,"RAPPROCHEMENT BANCAIRE",BH,bold=True,color="FFFFFF",align="center")
    ws.row_dimensions[row].height=22; row+=1
    for lbl_l,val_l,lbl_r,val_r in [("CLIENT",info.get('cl',''),"BANQUE",info.get('bq','')),
                                      ("N° COMPTE",info.get('co',''),"PÉRIODE",periode)]:
        cell(row,1,lbl_l,GY,bold=True); ws.merge_cells(f'B{row}:C{row}')
        cell(row,2,val_l,WH); ws.merge_cells(f'G{row}:H{row}')
        cell(row,6,lbl_r,GY,bold=True); cell(row,7,val_r,WH); row+=1
    ws.merge_cells(f'A{row}:E{row}'); cell(row,1,"RELEVE BANCAIRE",BH,bold=True,color="FFFFFF",align="center")
    ws.merge_cells(f'F{row}:J{row}'); cell(row,6,"JOURNAL BANQUE",BH,bold=True,color="FFFFFF",align="center"); row+=1
    for i,h in enumerate(['DATE','LIBELLE','N° PIECE','DEBIT','CREDIT','DATE','LIBELLE','N° PIECE','DEBIT','CREDIT'],1):
        al='right' if h in ('DEBIT','CREDIT') else ('left' if h=='LIBELLE' else 'center')
        cell(row,i,h,BH,bold=True,color="FFFFFF",align=al)
    row+=1
    rsd=e['rb_sol_d'] if e['rb_sol_d']>0 else None; rsc=e['rb_sol_c'] if e['rb_sol_c']>0 else None
    csd=e['cp_sol_d'] if e['cp_sol_d']>0 else None; csc=e['cp_sol_c'] if e['cp_sol_c']>0 else None
    for c in range(1,11): cell(row,c,fill_=BL if c<=5 else OL)
    cell(row,2,f"SOLDE AVANT RAPPROCHEMENT AU {periode}",BL,bold=True)
    for col,val,fl in [(4,rsd,BL),(5,rsc,BL)]:
        ws.cell(row=row,column=col,value=val).fill=fl
        ws.cell(row=row,column=col).alignment=Alignment(horizontal='right')
        ws.cell(row=row,column=col).font=Font(bold=True,size=10)
    cell(row,7,f"SOLDE AVANT RAPPROCHEMENT AU {periode}",OL,bold=True)
    for col,val,fl in [(9,csd,OL),(10,csc,OL)]:
        ws.cell(row=row,column=col,value=val).fill=fl
        ws.cell(row=row,column=col).alignment=Alignment(horizontal='right')
        ws.cell(row=row,column=col).font=Font(bold=True,size=10)
    for c in range(1,11): ws.cell(row=row,column=c).border=tb()
    row+=1

    for i in range(n_rows):
        if i < len(rows_left):
            item = rows_left[i]; r = item['row']; side = item['side']; is_rep = item['is_rep']
            rb_fl = RL if is_rep else BL
            lb = r['lib']
            if is_rep: lb = f'[↩ {r.get("carry_from","")}] {lb}'
            if side == 'gl_debit':
                rb_d = None; rb_c = r['debit'] if r['debit'] > 0 else None
            elif side == 'rb_debit':
                rb_d = r['debit'] if r['debit'] > 0 else None; rb_c = None
            else:  # gl_credit_courant
                rb_d = r['credit'] if r['credit'] > 0 else None; rb_c = None
            cell(row,1,r['date'],rb_fl,bold=is_rep); cell(row,2,lb,rb_fl,bold=is_rep)
            cell(row,3,r.get('piece',''),rb_fl,bold=is_rep)
            for col,val in [(4,rb_d),(5,rb_c)]:
                ws.cell(row=row,column=col,value=val).fill=rb_fl
                ws.cell(row=row,column=col).alignment=Alignment(horizontal='right')
                ws.cell(row=row,column=col).font=Font(bold=is_rep,size=10)
        else:
            for c in range(1,6): ws.cell(row=row,column=c).fill=BE; ws.cell(row=row,column=c).border=tb()

        if i < len(rows_right):
            pair2   = rows_right[i]
            d_item2 = pair2.get('debit_item')   # col 9 — DÉBIT journal
            c_item2 = pair2.get('credit_item')  # col 10 — CRÉDIT journal
            # Prendre la ligne principale pour date/libellé
            main2    = d_item2 if d_item2 else c_item2
            r2       = main2['row']; is_rep2 = main2['is_rep']
            cp_fl    = RL if is_rep2 else OL
            lb2      = r2['lib']
            if is_rep2: lb2 = f'[↩ {r2.get("carry_from","")}] {lb2}'
            cell(row,6,r2['date'],cp_fl,bold=is_rep2)
            cell(row,7,lb2,cp_fl,bold=is_rep2)
            cell(row,8,r2.get('piece',''),cp_fl,bold=is_rep2)
            # col 9 = DÉBIT journal (suspens RB crédit → débit journal)
            val_d9  = d_item2['row']['credit'] if d_item2 and d_item2['row']['credit'] > 0 else None
            # col 10 = CRÉDIT journal (GL crédits reportés)
            val_c10 = c_item2['row']['credit'] if c_item2 and c_item2['row']['credit'] > 0 else None
            for col,val in [(9, val_d9),(10, val_c10)]:
                ws.cell(row=row,column=col,value=val).fill=cp_fl
                ws.cell(row=row,column=col).alignment=Alignment(horizontal='right')
                ws.cell(row=row,column=col).font=Font(bold=is_rep2,size=10)
        else:
            for c in range(6,11): ws.cell(row=row,column=c).fill=OE; ws.cell(row=row,column=c).border=tb()
        for c in range(1,11): ws.cell(row=row,column=c).border=tb()
        row+=1

    for c in range(1,11):
        cl=ws.cell(row=row,column=c); cl.fill=ST; cl.font=Font(bold=True,color="FFFFFF",size=10); cl.border=tb()
    ws.cell(row=row,column=1,value="TOTAUX"); ws.cell(row=row,column=6,value="TOTAUX")
    for c,v in [(4,e['tot_d_rb']),(5,e['tot_c_rb']),(9,e['tot_d_cp']),(10,e['tot_c_cp'] if e['tot_c_cp'] else None)]:
        ws.cell(row=row,column=c,value=v).alignment=Alignment(horizontal='right')
    row+=1
    for c in range(1,11):
        cl=ws.cell(row=row,column=c); cl.fill=CY; cl.font=Font(bold=True,color="FFFFFF",size=10); cl.border=tb()
    ws.cell(row=row,column=2,value="SOLDES RAPPROCHES")
    ws.cell(row=row,column=4,value=e['sr_rb']).alignment=Alignment(horizontal='right')
    ws.cell(row=row,column=7,value="SOLDES RAPPROCHES")
    ws.cell(row=row,column=10,value=e['sr_cp']).alignment=Alignment(horizontal='right')
    row+=1
    for c in range(1,11): ws.cell(row=row,column=c).border=tb()
    ws.cell(row=row,column=5,value="Statut").fill=fill("E2EFDA")
    ws.cell(row=row,column=5).font=Font(bold=True,size=10)
    ws.cell(row=row,column=5).alignment=Alignment(horizontal='center')
    ws.merge_cells(f'F{row}:J{row}')
    cl=ws.cell(row=row,column=6,value='OK' if e['ok'] else 'A RAPPROCHER')
    cl.fill=GR if e['ok'] else RD; cl.font=Font(bold=True,color="FFFFFF",size=11)
    cl.alignment=Alignment(horizontal='center',vertical='center')
    for i,w in enumerate([13,42,12,15,15,13,42,12,15,15],1):
        ws.column_dimensions[get_column_letter(i)].width=w
    buf=BytesIO(); wb.save(buf); buf.seek(0)
    return buf.getvalue()

def importer_erb_fichier(uploaded_file, periode_source):
    try:
        uploaded_file.seek(0)
        wb = openpyxl.load_workbook(uploaded_file, data_only=True)
        ws = wb.active; rows = list(ws.iter_rows(values_only=True))
    except Exception as ex:
        return None, None, str(ex)
    header_row = None
    for i, row in enumerate(rows):
        vals = [str(v).strip().upper() if v else '' for v in row]
        if vals[0] in ('DATE','DATE ') and any('LIBELLE' in v for v in vals):
            header_row = i; break
    if header_row is None:
        return None, None, "Structure ERB non reconnue — ligne d'en-tête (DATE | LIBELLE) introuvable"
    def pn_raw(v):
        if v is None: return 0.0
        s = str(v).strip()
        if s.startswith('=') or s in ('','-','—'): return 0.0
        try: return abs(float(s.replace(' ','').replace('\xa0','').replace(',','.')))
        except: return 0.0
    def fmt_date(v):
        if v is None: return ''
        s = str(v).strip()
        if not s or s in ('nan','None',''): return ''
        try:
            n = float(s)
            if 20000 < n < 60000:
                import datetime
                d = datetime.date(1899, 12, 30) + datetime.timedelta(days=int(n))
                return d.strftime('%d/%m/%Y')
        except: pass
        try:
            d = pd.to_datetime(s, dayfirst=False if re.match(r'^\d{4}[-/]',s) else True, errors='coerce')
            return d.strftime('%d/%m/%Y') if d is not pd.NaT else s[:10]
        except: return s[:10]
    susp_rb = []; susp_cp = []
    for row in rows[header_row+1:]:
        if not any(v for v in row): continue
        lb1 = str(row[1] or '').strip().upper() if len(row)>1 else ''
        if any(k in lb1 for k in ('TOTAUX','SOLDES RAPPROCHES','SOLDE RAPPROCHE')): break
        rb_lib = str(row[1] or '').strip() if len(row)>1 else ''
        rb_d = pn_raw(row[3] if len(row)>3 else None)
        rb_c = pn_raw(row[4] if len(row)>4 else None)
        if rb_lib and 'SOLDE AVANT' not in rb_lib.upper() and (rb_d>0 or rb_c>0):
            susp_rb.append({'date':fmt_date(row[0]),'lib':rb_lib,
                'piece':str(row[2] or '').strip() if len(row)>2 else '',
                'debit':rb_d,'credit':rb_c,'matched':False,'match_id':None,'carry_from':periode_source})
        cp_lib = str(row[6] or '').strip() if len(row)>6 else ''
        cp_d = pn_raw(row[8] if len(row)>8 else None)
        cp_c = pn_raw(row[9] if len(row)>9 else None)
        if cp_lib and 'SOLDE AVANT' not in cp_lib.upper() and (cp_d>0 or cp_c>0):
            susp_cp.append({'date':fmt_date(row[5] if len(row)>5 else None),'lib':cp_lib,
                'piece':str(row[7] or '').strip() if len(row)>7 else '',
                'debit':cp_d,'credit':cp_c,'matched':False,'match_id':None,'carry_from':periode_source})
    return susp_rb, susp_cp, None

def page_import_erb():
    st.markdown('<div class="page-title">📂 Import ERB précédent</div>', unsafe_allow_html=True)
    st.markdown("<div class='page-sub'>Chargez le tableau ERB d'un mois antérieur pour régulariser ses suspens</div>", unsafe_allow_html=True)
    n_rb = len(st.session_state.carryover_rb); n_cp = len(st.session_state.carryover_cp)
    if n_rb>0 or n_cp>0:
        st.success(f"✅ Reports actuels chargés : **{n_rb}** suspens relevé · **{n_cp}** suspens GL")
        if st.button("🗑️ Effacer tous les reports", key="clear_carry_erb"):
            st.session_state.carryover_rb=[]; st.session_state.carryover_cp=[]
            st.session_state.regularisees=[]; st.rerun()
        st.markdown("---")
    periode_source = st.text_input("Période du tableau ERB importé", placeholder="Ex: 31 JUIN 2025",
        help="Ce libellé s'affichera en rouge dans le tableau du mois courant")
    uploaded_erb = st.file_uploader("Sélectionnez le fichier ERB (.xlsx)", type=['xlsx','xls'], key='uploader_erb_prev')
    if uploaded_erb and periode_source.strip():
        susp_rb, susp_cp, err = importer_erb_fichier(uploaded_erb, periode_source.strip())
        if err: st.error(f"❌ Erreur de lecture : {err}"); return
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(f"**📥 Suspens RELEVÉ extraits : {len(susp_rb)}**")
            if susp_rb:
                df_rb = pd.DataFrame(susp_rb)[['date','lib','debit','credit']]
                df_rb.columns = ['Date','Libellé','Débit','Crédit']
                st.dataframe(df_rb, use_container_width=True, hide_index=True, height=250)
            else: st.info("Aucun suspens côté Relevé")
        with col2:
            st.markdown(f"**📄 Suspens JOURNAL extraits : {len(susp_cp)}**")
            if susp_cp:
                df_cp = pd.DataFrame(susp_cp)[['date','lib','debit','credit']]
                df_cp.columns = ['Date','Libellé','Débit','Crédit']
                st.dataframe(df_cp, use_container_width=True, hide_index=True, height=250)
            else: st.info("Aucun suspens côté Journal")
        total = len(susp_rb)+len(susp_cp)
        if total == 0: st.warning("⚠️ Aucun suspens trouvé dans ce fichier ERB."); return
        if st.button(f"✅ Charger ces {total} suspens comme reports du mois précédent",
                     type="primary", use_container_width=True):
            # Remplacer complètement les carryover (pas d'extension partielle)
            # pour garantir que TOUS les suspens de l'ERB précédent sont présents
            st.session_state.carryover_rb = list(susp_rb)
            st.session_state.carryover_cp = list(susp_cp)
            st.session_state.rematch_needed = True
            n_rb_ok = st.session_state.rb_df is not None and len(st.session_state.rb_df) > 0
            n_cp_ok = st.session_state.cp_df is not None and len(st.session_state.cp_df) > 0
            if n_rb_ok and n_cp_ok:
                st.success(f"✅ **{len(susp_rb)}** suspens relevé + **{len(susp_cp)}** suspens GL chargés. Rapprochement en cours...")
            else:
                st.success(f"✅ **{len(susp_rb)}** suspens relevé + **{len(susp_cp)}** suspens GL chargés depuis **{periode_source}**.")
            st.session_state.page = "rapprochement"; st.rerun()
    elif uploaded_erb and not periode_source.strip():
        st.warning("⚠️ Saisissez d'abord la période du tableau ERB importé.")

# ───────────────────────────────────────────────────────────────────────
# PAGE IMPORT (relevé / comptabilité)  — identique v5.5
# ───────────────────────────────────────────────────────────────────────
def page_import(side):
    label="Relevé bancaire" if side=='rb' else "Écritures comptables"
    icon ="📥" if side=='rb' else "📄"
    st.markdown(f'<div class="page-title">{icon} Import — {label}</div>', unsafe_allow_html=True)
    st.markdown(f'<div class="page-sub">Chargez votre fichier Excel ou CSV</div>', unsafe_allow_html=True)
    uploaded=st.file_uploader("Cliquez ici pour choisir votre fichier",type=['xlsx','xls','csv','pdf'],
                               key=f'uploader_{side}',label_visibility='visible')
    if not uploaded: return
    if uploaded.size > 20 * 1024 * 1024:
        st.error(f"⚠️ Fichier trop volumineux ({uploaded.size/1024/1024:.1f} MB). Limite : 20 MB.")
        return
    is_pdf = uploaded.name.lower().endswith('.pdf')
    if is_pdf:
        st.info("📄 PDF détecté — extraction automatique du tableau en cours...")
    df_raw=load_file(uploaded)
    if df_raw is None:
        if is_pdf:
            uploaded.seek(0)
            _, pdf_err = _extract_pdf_tables(uploaded)
            if pdf_err and "scanné" in pdf_err:
                st.error("❌ **PDF scanné (image)** — Tentative de lecture par IA...")
                # Essayer OCR via Claude
                has_key = False
                try:
                    has_key = bool(st.secrets.get("ANTHROPIC_API_KEY",""))
                except Exception:
                    pass
                if has_key:
                    with st.spinner("🤖 L'IA lit votre relevé scanné... (peut prendre 15-30 secondes)"):
                        uploaded.seek(0)
                        df_ocr, ocr_err = _extract_pdf_scanned_via_claude(uploaded)
                    if df_ocr is not None and not df_ocr.empty:
                        st.success(f"✅ IA a extrait **{len(df_ocr)} transactions** du relevé scanné !")
                        # Mapper directement les colonnes détectées
                        rename_map = {"libelle":"Libellé","date":"Date","piece":"Réf.","debit":"Débit","credit":"Crédit"}
                        df_show = df_ocr.rename(columns=rename_map)
                        st.dataframe(df_show, use_container_width=True, hide_index=True, height=250)
                        # Convertir en format attendu par l'app
                        df_proc = pd.DataFrame({
                            "date":    df_ocr.get("date",""),
                            "lib":     df_ocr.get("libelle",""),
                            "piece":   df_ocr.get("piece",""),
                            "debit":   pd.to_numeric(df_ocr.get("debit",0), errors="coerce").fillna(0.0),
                            "credit":  pd.to_numeric(df_ocr.get("credit",0), errors="coerce").fillna(0.0),
                            "matched": False,
                            "match_id": None,
                        })
                        # Filtrer lignes vides
                        df_proc = df_proc[(df_proc["debit"]>0)|(df_proc["credit"]>0)].reset_index(drop=True)
                        if side == "rb":
                            st.session_state.rb_df = df_proc
                            st.session_state.rb_map = {"date":"date","lib":"libelle","debit":"debit","credit":"credit"}
                        else:
                            st.session_state.cp_df = df_proc
                            st.session_state.cp_map = {"date":"date","lib":"libelle","debit":"debit","credit":"credit"}
                        st.success(f"✅ {len(df_proc)} lignes importées depuis le relevé scanné.")
                        st.rerun()
                        return
                    else:
                        st.error(f"❌ Lecture IA échouée : {ocr_err}")
                        st.warning("Essayez : ilovepdf.com → PDF to Excel, puis importez l'Excel.")
                else:
                    st.warning("""**PDF scanné — Solutions :**
- **Recommandé** : Ajoutez `ANTHROPIC_API_KEY` dans `.streamlit/secrets.toml` pour lecture automatique par IA
- **Méthode manuelle** : ilovepdf.com → "PDF to Excel" (gratuit), puis importez l'Excel
- **Portail Ecobank** : exportez directement en PDF numérique ou Excel""")
            else:
                st.error(f"❌ Extraction PDF échouée : {pdf_err}")
                st.info("💡 Convertissez en Excel via smallpdf.com ou ilovepdf.com.")
        else:
            st.error("❌ Impossible de lire le fichier. Vérifiez qu'il s'agit d'un Excel (.xlsx/.xls), CSV ou PDF valide.")
        return
    hi=detect_header_row(df_raw)
    headers=df_raw.iloc[hi].fillna('').astype(str).str.strip().tolist()
    df_data=df_raw.iloc[hi+1:].copy(); df_data.columns=headers
    df_data=df_data[df_data.apply(lambda r: r.astype(str).str.strip().ne('').any(),axis=1)].reset_index(drop=True)
    st.success(f"✅ {uploaded.name} — {len(df_data)} lignes détectées")
    if side=='rb': st.session_state.rb_raw=df_data; st.session_state.rb_hd=headers
    else:          st.session_state.cp_raw=df_data; st.session_state.cp_hd=headers
    st.markdown("##### Correspondance des colonnes")
    avail=['— ignorer —']+headers; user_map={}
    cols_ui=st.columns(4)
    for idx,field in enumerate(['date','lib','piece','debit','credit','montant','sens']):
        with cols_ui[idx%4]:
            guessed=guess_col(headers,field)
            def_idx=avail.index(guessed) if guessed in avail else 0
            sel=st.selectbox(FL[field],avail,index=def_idx,key=f'ms_{side}_{field}')
            if sel!='— ignorer —': user_map[field]=sel
    if st.button("✅ Appliquer et importer",type="primary",key=f'btn_apply_{side}'):
        # Validation du mapping minimum requis
        has_amount = 'debit' in user_map or 'credit' in user_map or 'montant' in user_map
        if not has_amount:
            st.error("❌ Mappez au minimum les colonnes **Débit** et **Crédit** (ou **Montant**) avant d'importer.")
            st.stop()
        df_proc,solde_auto=apply_map(df_data,user_map)
        if df_proc.empty: st.error("❌ Aucune ligne valide extraite. Vérifiez le mapping des colonnes Débit/Crédit."); return
        if side=='rb':
            st.session_state.rb_df=df_proc; st.session_state.rb_map=user_map
            if solde_auto: st.session_state.rb_solde_auto=solde_auto
        else:
            st.session_state.cp_df=df_proc; st.session_state.cp_map=user_map
            if solde_auto: st.session_state.cp_solde_auto=solde_auto
        rb_ready=st.session_state.rb_df is not None and len(st.session_state.rb_df)>0
        cp_ready=st.session_state.cp_df is not None and len(st.session_state.cp_df)>0
        if rb_ready and cp_ready:
            inv = st.session_state.get('inv_tog', True)
            rb_full = get_rb_with_carryover(); cp_full = get_cp_with_carryover()
            rb_full = rb_full.copy(); cp_full = cp_full.copy()
            rb_full['matched']=False; rb_full['match_id']=None
            cp_full['matched']=False; cp_full['match_id']=None
            rb_m, cp_m, cnt = auto_match(rb_full, cp_full, inversion=inv)
            st.session_state.regularisees = extraire_regularisees(rb_m, cp_m)
            if 'carried' in rb_m.columns:
                st.session_state.rb_df = rb_m[~rb_m['carried'].fillna(False).astype(bool)].drop(columns=['carried'], errors='ignore').copy()
                carried_rb = rb_m[rb_m['carried'].fillna(False).astype(bool)].copy()
                non_reg = carried_rb[~carried_rb['matched']].drop(columns=['carried'], errors='ignore').copy()
                if len(non_reg) > 0:
                    st.session_state.rb_df = pd.concat([st.session_state.rb_df, non_reg], ignore_index=True)
                st.session_state.carryover_rb = []
            else:
                st.session_state.rb_df = rb_m
            if 'carried' in cp_m.columns:
                st.session_state.cp_df = cp_m[~cp_m['carried'].fillna(False).astype(bool)].drop(columns=['carried'], errors='ignore').copy()
                carried_cp = cp_m[cp_m['carried'].fillna(False).astype(bool)].to_dict('records')
                st.session_state.carryover_cp = [r for r in carried_cp if not r.get('matched')]
            else:
                st.session_state.cp_df = cp_m
            st.session_state.rematch_needed = False
            n_reg = len(st.session_state.regularisees)
            msg = f"✅ {len(df_proc)} lignes importées · **{cnt} rapprochements** effectués"
            if n_reg > 0: msg += f" dont **{n_reg} régularisation(s)** de suspens reportés"
            st.success(msg)
        else:
            st.success(f"✅ {len(df_proc)} lignes importées")
        st.dataframe(df_proc[['date','lib','piece','debit','credit']].head(50)
            .rename(columns={'date':'Date','lib':'Libellé','piece':'Réf.','debit':'Débit','credit':'Crédit'}),
            use_container_width=True,hide_index=True)

# ───────────────────────────────────────────────────────────────────────
# PAGE MES ERB SAUVEGARDÉS
# ───────────────────────────────────────────────────────────────────────
def page_mes_erb():
    st.markdown('<div class="page-title">📚 Mes ERB sauvegardés</div>', unsafe_allow_html=True)
    st.markdown("<div class='page-sub'>Retrouvez et téléchargez tous vos tableaux ERB archivés</div>", unsafe_allow_html=True)

    licence_code = st.session_state.get('_licence_code','')
    sb = get_supabase()

    if sb is None:
        err_detail = st.session_state.get("_supabase_error", "")
        msg = "**Supabase non configuré** — La sauvegarde cloud n'est pas encore activée."
        if err_detail:
            msg += f"\n\n**Erreur détaillée :** `{err_detail}`"
        st.warning(msg)
        with st.expander("SQL à exécuter dans Supabase", expanded=True):
            st.code("""
CREATE TABLE erb_historique (
    id            BIGSERIAL PRIMARY KEY,
    entreprise    TEXT NOT NULL,
    licence_code  TEXT NOT NULL,
    periode       TEXT NOT NULL,
    date_creation TIMESTAMPTZ DEFAULT NOW(),
    sr_rb         NUMERIC,
    sr_cp         NUMERIC,
    equilibre     BOOLEAN,
    xlsx_base64   TEXT
);
CREATE INDEX idx_erb_licence ON erb_historique(licence_code);
""", language="sql")
        return

    with st.spinner("Chargement de votre historique..."):
        historique = charger_historique_erb(licence_code)

    if not historique:
        st.info("Aucun ERB sauvegardé pour le moment. Générez un tableau ERB pour qu'il soit automatiquement archivé.")
        return

    n_ok = sum(1 for h in historique if h.get('equilibre'))
    c1, c2, c3 = st.columns(3)
    c1.metric("ERB archivés", len(historique))
    c2.metric("Equilibres", n_ok)
    c3.metric("A rapprocher", len(historique) - n_ok)
    st.markdown("---")

    for h in historique:
        date_str  = h.get('date_creation','')[:10] if h.get('date_creation') else ''
        periode   = h.get('periode','')
        sr_rb     = h.get('sr_rb', 0) or 0
        sr_cp     = h.get('sr_cp', 0) or 0
        equilibre = h.get('equilibre', False)
        erb_id    = h.get('id')
        status_color = "#1B5E20" if equilibre else "#E65100"
        status_label = "Equilibre" if equilibre else "A rapprocher"

        col_info, col_dl = st.columns([4, 1])
        with col_info:
            st.markdown(f"""
            <div style="background:rgba(255,255,255,0.05);border:1px solid rgba(100,160,255,0.15);
                        border-radius:10px;padding:12px 16px;margin-bottom:8px">
              <div style="display:flex;justify-content:space-between;align-items:center">
                <div>
                  <span style="font-size:15px;font-weight:700;color:#fff">📅 {periode}</span>
                  <span style="font-size:11px;color:#607D8B;margin-left:12px">Archivé le {date_str}</span>
                </div>
                <div style="background:{status_color};border-radius:6px;padding:3px 10px;
                            font-size:12px;font-weight:700;color:#fff">{status_label}</div>
              </div>
              <div style="margin-top:8px;font-size:12px;color:#90CAF9">
                SR Releve : <b>{sr_rb:,.0f}</b> &nbsp;|&nbsp; SR Comptabilite : <b>{sr_cp:,.0f}</b>
              </div>
            </div>""", unsafe_allow_html=True)
        with col_dl:
            st.markdown("<div style='margin-top:12px'>", unsafe_allow_html=True)
            if st.button("📥", key=f"dl_{erb_id}", use_container_width=True, help="Télécharger cet ERB"):
                with st.spinner("Chargement..."):
                    xlsx_bytes, fname = telecharger_erb_sauvegarde(erb_id)
                if xlsx_bytes:
                    st.download_button(
                        label="Télécharger",
                        data=xlsx_bytes,
                        file_name=fname,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_file_{erb_id}",
                    )
                else:
                    st.error("Erreur")
            st.markdown("</div>", unsafe_allow_html=True)

# ═══ SIDEBAR ══════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div class="logo-box">🏦 ERB Manager</div>', unsafe_allow_html=True)
    entreprise = st.session_state.get('entreprise','')
    user_email = st.session_state.get('_user_email','')
    if entreprise:
        st.markdown(f'''
        <div style="background:rgba(255,255,255,0.08);border-radius:8px;padding:8px 10px;margin-bottom:8px">
          <div style="font-size:12px;color:rgba(163,210,230,0.95);font-weight:700">🏢 {entreprise}</div>
          {f'<div style="font-size:10px;color:rgba(163,210,230,0.6);margin-top:2px">👤 {user_email}</div>' if user_email else ''}
        </div>
        ''', unsafe_allow_html=True)
    st.markdown('<div class="logo-sub">Rapprochement bancaire</div>', unsafe_allow_html=True)
    rb_ok=st.session_state.rb_df is not None and len(st.session_state.rb_df)>0
    cp_ok=st.session_state.cp_df is not None and len(st.session_state.cp_df)>0
    for key,icon,label in [("dashboard","📊","Tableau de bord"),("import_rb","📥","Relevé bancaire"),
                             ("import_cp","📄","Écritures comptables"),
                             ("import_erb_prev","📂","ERB précédent"),
                             ("rapprochement","🔗","Rapprochement"),
                             ("mes_erb","📚","Mes ERB sauvegardés")]:
        active=st.session_state.page==key
        if st.button(f"{icon}  {label}",key=f"nav_{key}",use_container_width=True,
                     type="primary" if active else "secondary"):
            st.session_state.page=key; st.rerun()
    st.divider()
    st.markdown(f"**Relevé :** {'✅ '+str(len(st.session_state.rb_df))+' lignes' if rb_ok else '⏳ en attente'}")
    st.markdown(f"**Compta :** {'✅ '+str(len(st.session_state.cp_df))+' lignes' if cp_ok else '⏳ en attente'}")
    if rb_ok and cp_ok:
        st.markdown(f"**Rapprochés :** ✅ {int(st.session_state.rb_df['matched'].sum())}")
    reg=st.session_state.get('regularisees',[])
    if reg: st.markdown(f"**✅ Régularisés :** {len(reg)} suspens soldés")
    st.divider()
    if st.button("🗑️ Réinitialiser tout", key="nav_reset_all", use_container_width=True,
                 help="Vider toute la session et repartir de zéro"):
        licence_bak = {k: st.session_state[k] for k in ['_licence_ok','_licence_code','entreprise'] if k in st.session_state}
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.session_state.update(licence_bak)
        st.rerun()
    if st.button("🚪 Se déconnecter", key="nav_logout", use_container_width=True,
                 help="Fermer la session et revenir à l'écran de licence"):
        for k in list(st.session_state.keys()):
            del st.session_state[k]
        st.rerun()
    st.caption("ERB v5.17")

# ═══ ROUTING ══════════════════════════════════════════════════════════
page=st.session_state.page

if page=="dashboard":
    st.markdown('<div class="page-title">📊 Tableau de bord</div>', unsafe_allow_html=True)
    st.markdown("<div class='page-sub'>Vue d'ensemble du rapprochement</div>", unsafe_allow_html=True)
    rb_ok=st.session_state.rb_df is not None and len(st.session_state.rb_df)>0
    cp_ok=st.session_state.cp_df is not None and len(st.session_state.cp_df)>0
    rb_n=len(st.session_state.rb_df) if rb_ok else 0
    cp_n=len(st.session_state.cp_df) if cp_ok else 0
    mat_n=int(st.session_state.rb_df['matched'].sum()) if rb_ok else 0
    c1,c2,c3,c4=st.columns(4)
    c1.metric("📥 Relevé",rb_n,"lignes importées"); c2.metric("📄 Comptabilité",cp_n,"lignes importées")
    c3.metric("✅ Rapprochées",mat_n)
    c4.metric("📈 Taux",f"{int(mat_n/max(rb_n,cp_n,1)*100)} %" if (rb_ok or cp_ok) else "—")
    st.divider()
    col1,col2=st.columns(2)
    with col1:
        st.markdown("**État des imports**")
        st.markdown(f"{'✅' if rb_ok else '❌'} Relevé bancaire")
        st.markdown(f"{'✅' if cp_ok else '❌'} Écritures comptables")
        st.markdown(f"{'✅' if rb_ok and cp_ok else '⏳'} Rapprochement")
    with col2:
        st.markdown("**Actions rapides**")
        if st.button("📥 Importer le relevé bancaire",use_container_width=True):
            st.session_state.page="import_rb"; st.rerun()
        if st.button("📄 Importer les écritures comptables",use_container_width=True):
            st.session_state.page="import_cp"; st.rerun()
        if st.button("🔗 Lancer le rapprochement",type="primary",use_container_width=True):
            st.session_state.page="rapprochement"; st.rerun()

elif page=="import_rb":    page_import("rb")
elif page=="import_cp":    page_import("cp")
elif page=="import_erb_prev": page_import_erb()
elif page=="mes_erb":          page_mes_erb()

elif page=="rapprochement":
    st.markdown('<div class="page-title">🔗 Rapprochement bancaire</div>', unsafe_allow_html=True)
    st.markdown('<div class="page-sub">Tableau ERB — correspondances et calcul des écarts</div>', unsafe_allow_html=True)

    # ── Auto-rematch si nécessaire ────────────────────────────────────
    _rb_ok = st.session_state.rb_df is not None and len(st.session_state.rb_df) > 0
    _cp_ok = st.session_state.cp_df is not None and len(st.session_state.cp_df) > 0
    _need  = st.session_state.get('rematch_needed', False)
    _has_carry = len(st.session_state.carryover_rb) > 0 or len(st.session_state.carryover_cp) > 0

    _needs_passe1 = False
    if _rb_ok and not _has_carry and not _need:
        rb_tmp = st.session_state.rb_df
        if 'carry_from' in rb_tmp.columns:
            has_rep = (rb_tmp['carry_from'].notna() &
                       ~rb_tmp['carry_from'].astype(str).isin(['','nan','None','False']) &
                       ~rb_tmp['matched']).any()
            has_cur = ((rb_tmp['carry_from'].isna() |
                        rb_tmp['carry_from'].astype(str).isin(['','nan','None','False'])) &
                       ~rb_tmp['matched']).any()
            if has_rep and has_cur: _needs_passe1 = True

    if _rb_ok and _cp_ok and (_need or _has_carry or _needs_passe1):
        # Anti-boucle : si déjà en train de rematching, ne pas relancer
        if not st.session_state.get('_rematching', False):
            st.session_state['_rematching'] = True
            st.session_state.rematch_needed = False

            with st.spinner("🔄 Rapprochement en cours..."):
                rb_tmp = st.session_state.rb_df
                if rb_tmp is not None and 'carry_from' in rb_tmp.columns:
                    mask_rep = (rb_tmp['carry_from'].notna() &
                                ~rb_tmp['carry_from'].astype(str).isin(['','nan','None','False']))
                    if mask_rep.any():
                        reports_back = rb_tmp[mask_rep].copy()
                        reports_back['carried'] = True; reports_back['matched'] = False; reports_back['match_id'] = None
                        st.session_state.rb_df = rb_tmp[~mask_rep].copy()
                        existing_libs = {r['lib'] for r in st.session_state.carryover_rb}
                        nouveaux = [r for r in reports_back.to_dict('records') if r['lib'] not in existing_libs]
                        st.session_state.carryover_rb.extend(nouveaux)
                rb_full = get_rb_with_carryover(); cp_full = get_cp_with_carryover()
                rb_full = rb_full.copy(); cp_full = cp_full.copy()
                rb_full['matched'] = False; rb_full['match_id'] = None
                cp_full['matched'] = False; cp_full['match_id'] = None
                st.session_state.mc = 0
                rb_m, cp_m, cnt = auto_match(rb_full, cp_full, inversion=st.session_state.get('inv_tog', True))
                st.session_state.regularisees = extraire_regularisees(rb_m, cp_m)
                n_reg = len(st.session_state.regularisees)
                if 'carried' in rb_m.columns:
                    st.session_state.rb_df = rb_m[~rb_m['carried'].fillna(False).astype(bool)].drop(columns=['carried'], errors='ignore').copy()
                    carried_rb = rb_m[rb_m['carried'].fillna(False).astype(bool)].copy()
                    non_reg = carried_rb[~carried_rb['matched'].fillna(False).astype(bool)].drop(columns=['carried'], errors='ignore').copy()
                    if len(non_reg) > 0:
                        st.session_state.rb_df = pd.concat([st.session_state.rb_df, non_reg], ignore_index=True)
                    st.session_state.carryover_rb = []
                else:
                    st.session_state.rb_df = rb_m
                if 'carried' in cp_m.columns:
                    st.session_state.cp_df = cp_m[~cp_m['carried'].fillna(False).astype(bool)].drop(columns=['carried'], errors='ignore').copy()
                    carried_cp = cp_m[cp_m['carried'].fillna(False).astype(bool)].to_dict('records')
                    st.session_state.carryover_cp = [r for r in carried_cp if not r.get('matched')]
                else:
                    st.session_state.cp_df = cp_m
                # Invalider le cache xlsx
                for k in list(st.session_state.keys()):
                    if k.startswith('_xlsx_'): del st.session_state[k]
                # Stocker le message de succès et libérer le flag anti-boucle
                msg = f"🔄 Rapprochement relancé : **{cnt}** correspondances"
                if n_reg > 0: msg += f" dont **{n_reg} régularisation(s)** du mois précédent"
                st.session_state['_rematch_msg'] = msg
            st.session_state['_rematching'] = False
            st.rerun()

    # Afficher le message de succès du rematch s'il existe
    if st.session_state.get('_rematch_msg'):
        st.success(st.session_state.pop('_rematch_msg'))

    # Historique
    if st.session_state.historique:
        with st.expander(f"📅 Historique — {len(st.session_state.historique)} mois clôturé(s)",expanded=False):
            for h in st.session_state.historique:
                st.markdown(f"**{h['periode']}** — {h['n_rb']} suspens RB · {h['n_cp']} suspens GL reportés")
                if h.get('erb_html'):
                    with st.expander(f"📋 Voir le tableau ERB — {h['periode']}",expanded=False):
                        st.markdown(f'<div style="overflow-x:auto">{h["erb_html"]}</div>', unsafe_allow_html=True)

    # Régularisations
    reg=st.session_state.get('regularisees',[])
    if reg:
        with st.expander(f"✅ {len(reg)} opération(s) du mois précédent régularisée(s) ce mois",expanded=False):
            for r in reg:
                montant=r['debit'] if r['debit']>0 else r['credit']
                st.markdown(f"**{r['type']}** _(reporté de {r['from']})_  \n"
                            f"📌 `{r['date']}` — {r['lib']} — **{fmt_fr(montant)}**  \n"
                            f"↔ Soldé par : `{r['partner_date']}` — {r['partner_lib']}")

    # Reports
    n_carry_rb=len(st.session_state.carryover_rb); n_carry_cp=len(st.session_state.carryover_cp)
    if n_carry_rb>0 or n_carry_cp>0:
        st.info(f"🔁 **Report du mois précédent :** {n_carry_rb} suspens relevé · {n_carry_cp} suspens GL")
        if st.button("🗑️ Effacer les reports",key="clear_carry"):
            st.session_state.carryover_rb=[]; st.session_state.carryover_cp=[]
            st.session_state.regularisees=[]; st.rerun()

    # ── Informations — persistance via on_change ──────────────────────
    def _save_srb(): st.session_state['s_rb'] = float(st.session_state['_inp_s_rb'])
    def _save_scp(): st.session_state['s_cp'] = float(st.session_state['_inp_s_cp'])

    rb_auto=st.session_state.rb_solde_auto; cp_auto=st.session_state.cp_solde_auto
    if rb_auto and float(st.session_state.get('s_rb',0.0))==0.0:
        st.session_state['s_rb'] = float(rb_auto[1])
    if cp_auto and float(st.session_state.get('s_cp',0.0))==0.0:
        st.session_state['s_cp'] = float(cp_auto[1])

    with st.expander("ℹ️ Informations",expanded=True):
        c1,c2,c3=st.columns(3)
        with c1:
            st.text_input("Client",   key='inf_cl', placeholder="Nom du client")
            st.text_input("Banque",   key='inf_bq', placeholder="Nom de la banque")
        with c2:
            st.text_input("N° compte",key='inf_co', placeholder="N° de compte")
            st.text_input("Période",  key='inf_pe', placeholder="Ex: 31 Juillet 2025")
        with c3:
            st.number_input("Solde final Relevé (créditeur ≥ 0 / débiteur < 0)",
                value=float(st.session_state.get('s_rb',0.0)), step=1.0, format="%.0f",
                key='_inp_s_rb', on_change=_save_srb)
            st.number_input("Solde final Comptabilité (débiteur ≥ 0 / créditeur < 0)",
                value=float(st.session_state.get('s_cp',0.0)), step=1.0, format="%.0f",
                key='_inp_s_cp', on_change=_save_scp)

    # Lire les valeurs persistées
    client  = st.session_state.get('inf_cl','')
    banque  = st.session_state.get('inf_bq','')
    compte  = st.session_state.get('inf_co','')
    periode = st.session_state.get('inf_pe','')
    s_rb    = float(st.session_state.get('s_rb',0.0))
    s_cp    = float(st.session_state.get('s_cp',0.0))

    inversion=st.toggle("🔄 Inversion D/C activée — Débit relevé = Crédit comptabilité",value=True,key='inv_tog')
    info={'cl':client,'bq':banque,'co':compte,'pe':periode,'s_rb':s_rb,'s_cp':s_cp}

    ac1,ac2,ac3,ac4,ac5=st.columns(5)
    with ac1:
        if st.button("🪄 Rapprochement auto",type="primary",use_container_width=True):
            rb=get_rb_with_carryover(); cp=get_cp_with_carryover()
            if rb is not None and len(rb)>0 and cp is not None and len(cp)>0:
                rb=rb.copy(); cp=cp.copy()
                rb['matched']=False; rb['match_id']=None
                cp['matched']=False; cp['match_id']=None
                rb_m,cp_m,cnt=auto_match(rb,cp,inversion)
                st.session_state.regularisees=extraire_regularisees(rb_m,cp_m)
                if 'carried' in rb_m.columns:
                    st.session_state.rb_df = rb_m[~rb_m['carried'].fillna(False).astype(bool)].drop(columns=['carried'], errors='ignore').copy()
                    carried_rb = rb_m[rb_m['carried'].fillna(False).astype(bool)].copy()
                    non_reg = carried_rb[~carried_rb['matched']].drop(columns=['carried'], errors='ignore').copy()
                    if len(non_reg) > 0:
                        st.session_state.rb_df = pd.concat([st.session_state.rb_df, non_reg], ignore_index=True)
                    st.session_state.carryover_rb = []
                else:
                    st.session_state.rb_df=rb_m
                if 'carried' in cp_m.columns:
                    st.session_state.cp_df=cp_m[~cp_m['carried'].fillna(False).astype(bool)].drop(columns=['carried'],errors='ignore').copy()
                    carried_cp=cp_m[cp_m['carried'].fillna(False).astype(bool)].to_dict('records')
                    st.session_state.carryover_cp=[r for r in carried_cp if not r.get('matched')]
                else:
                    st.session_state.cp_df=cp_m
                n_reg=len(st.session_state.regularisees)
                msg=f"✅ {cnt} rapprochement(s) effectué(s)"
                if n_reg>0: msg+=f" dont **{n_reg} régularisation(s)** de suspens reportés"
                # Invalider le cache xlsx pour forcer régénération
                for k in list(st.session_state.keys()):
                    if k.startswith('_xlsx_'): del st.session_state[k]
                st.success(msg); st.rerun()
            else:
                st.warning("Importez d'abord les deux fichiers.")
    with ac2:
        if st.button("🔗 Lier la sélection",use_container_width=True):
            st.info("Utilisez les cases à cocher dans les tableaux ci-dessous.")
    with ac3:
        if st.button("🔄 Réinitialiser",use_container_width=True):
            for side in ('rb_df','cp_df'):
                df=st.session_state[side]
                if df is not None: df['matched']=False; df['match_id']=None; st.session_state[side]=df
            st.session_state.mc=0
            for k in list(st.session_state.keys()):
                if k.startswith('_xlsx_'): del st.session_state[k]
            st.success("Liens réinitialisés."); st.rerun()
    with ac4:
        if st.button("➕ Ligne RB",use_container_width=True):
            if st.session_state.rb_df is not None:
                nr=pd.DataFrame([{'date':'','lib':'','piece':'','debit':0.0,'credit':0.0,'matched':False,'match_id':None}])
                st.session_state.rb_df=pd.concat([st.session_state.rb_df,nr],ignore_index=True); st.rerun()
    with ac5:
        if st.button("➕ Ligne CPTA",use_container_width=True):
            if st.session_state.cp_df is not None:
                nr=pd.DataFrame([{'date':'','lib':'','piece':'','debit':0.0,'credit':0.0,'matched':False,'match_id':None}])
                st.session_state.cp_df=pd.concat([st.session_state.cp_df,nr],ignore_index=True); st.rerun()

    rb_display=get_rb_with_carryover(); cp_display=get_cp_with_carryover()
    if rb_display is not None or cp_display is not None:
        st.markdown("---")
        t1,t2=st.columns(2)
        with t1:
            st.markdown("**📥 Relevé bancaire**")
            if rb_display is not None and len(rb_display)>0:
                cols=['date','lib','piece','debit','credit','matched']
                if 'carried' in rb_display.columns: cols.append('carried')
                df_show=rb_display[cols].copy()
                rename={'date':'Date','lib':'Libellé','piece':'Réf.','debit':'Débit','credit':'Crédit','matched':'✅','carried':'🔁'}
                df_show.columns=[rename.get(c,c) for c in df_show.columns]
                st.dataframe(df_show,use_container_width=True,height=320,hide_index=True)
        with t2:
            st.markdown("**📄 Journal comptable**")
            if cp_display is not None and len(cp_display)>0:
                cols=['date','lib','piece','debit','credit','matched']
                if 'carried' in cp_display.columns: cols.append('carried')
                df_show=cp_display[cols].copy()
                rename={'date':'Date','lib':'Libellé','piece':'Réf.','debit':'Débit','credit':'Crédit','matched':'✅','carried':'🔁'}
                df_show.columns=[rename.get(c,c) for c in df_show.columns]
                st.dataframe(df_show,use_container_width=True,height=320,hide_index=True)

    rb_for_erb=get_rb_with_carryover(); cp_for_erb=get_cp_with_carryover()
    if rb_for_erb is not None and len(rb_for_erb)>0 and cp_for_erb is not None and len(cp_for_erb)>0:
        rb_calc = rb_for_erb.drop(columns=['carried'], errors='ignore')
        cp_calc = cp_for_erb.drop(columns=['carried'], errors='ignore')
        e=calc_erb(rb_calc,cp_calc,s_rb,s_cp)

        st.markdown("---"); st.markdown("##### Calcul des écarts")
        m1,m2,m3=st.columns(3)
        m1.metric("Solde rapproché Relevé",    f"{e['sr_rb']:,.0f}")
        m2.metric("Solde rapproché Comptabilité",f"{e['sr_cp']:,.0f}")
        m3.metric("Statut","OK ✅" if e['ok'] else f"ÉCART ❌  ({e['chk']:,.0f})")
        if e['ok']: st.success("✅ Rapprochement équilibré — SR_RB = SR_CP")
        else: st.warning(f"⚠️ Écart de {e['chk']:,.0f} — {len(e['susp_rb'])} suspens relevé · {len(e['susp_cp'])} suspens GL")

        st.markdown("---"); st.markdown("##### 📋 Tableau ERB — Aperçu avant export")
        missing_rb=(s_rb==0); missing_cp=(s_cp==0)
        all_unmatched=len(rb_calc)>0 and not rb_calc['matched'].any()
        warn_items = []
        if missing_rb:    warn_items.append("- Saisissez le **Solde final du Relevé bancaire**")
        if missing_cp:    warn_items.append("- Saisissez le **Solde final de la Comptabilité**")
        if all_unmatched: warn_items.append("- Lancez le **🪄 Rapprochement automatique**")
        if warn_items:
            st.warning("⚠️ **Tableau ERB incomplet :**\n" + "\n".join(warn_items))

        # Toujours afficher le tableau (même incomplet) pour prévisualisation
        erb_html=build_erb_html(e,info)
        st.markdown(f'<div style="overflow-x:auto;background:#fff;border-radius:10px;padding:12px;box-shadow:0 4px 20px rgba(0,0,0,0.3)">{erb_html}</div>', unsafe_allow_html=True)

        n_rep_nonreg = sum(1 for _,r in e['susp_rb_debit'].iterrows() if _is_report(r.to_dict()))
        if n_rep_nonreg>0:
            st.markdown(f'<div style="font-size:11px;color:#C00;margin-top:4px">🟥 {n_rep_nonreg} suspens reporté(s) non régularisé(s) — apparaissent en rouge avec [↩ Mois d\'origine]</div>', unsafe_allow_html=True)

        st.markdown("")
        # Générer et stocker le fichier Excel en session_state
        # pour éviter qu'il soit recalculé/perdu lors du clic sur le bouton
        fname = f"ERB_{periode.replace(' ','_') if periode.strip() else 'export'}.xlsx"
        # Clé de cache incluant les soldes pour forcer régénération si soldes changent
        _cache_key = f"_xlsx_{fname}_{int(s_rb)}_{int(s_cp)}"
        if _cache_key not in st.session_state or st.session_state[_cache_key] is None:
            # Invalider anciens caches pour cette période
            for _k in list(st.session_state.keys()):
                if _k.startswith(f"_xlsx_{fname.replace('.xlsx','')}"): 
                    del st.session_state[_k]
            try:
                st.session_state[_cache_key] = export_xlsx(e, info)
            except Exception as _ex:
                st.error(f"⚠️ Erreur génération Excel : {_ex}")
                st.session_state[_cache_key] = None

        col_dl,col_cl=st.columns([2,1])
        with col_dl:
            if st.session_state.get(_cache_key):
                # Sauvegarder dans Supabase si pas encore fait pour cette période
                _save_key = f"_saved_{fname}_{int(s_rb)}_{int(s_cp)}"
                if not st.session_state.get(_save_key) and periode.strip():
                    ok_save, err_save = sauvegarder_erb(
                        entreprise    = st.session_state.get('entreprise',''),
                        licence_code  = st.session_state.get('_licence_code',''),
                        periode       = periode,
                        xlsx_bytes    = st.session_state[_cache_key],
                        sr_rb         = e['sr_rb'],
                        sr_cp         = e['sr_cp'],
                        ok            = e['ok'],
                    )
                    if ok_save:
                        st.session_state[_save_key] = True
                    # Echec silencieux (Supabase non configuré) — ne pas bloquer l'UX

                st.download_button(
                    label="📥 Télécharger le tableau ERB (.xlsx)",
                    data=st.session_state[_cache_key],
                    file_name=fname,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True)
            else:
                st.button("📥 Télécharger le tableau ERB (.xlsx)",
                          disabled=True, use_container_width=True,
                          help="Erreur de génération — vérifiez les données")
        with col_cl:
            if st.button("📅 Clôturer ce mois et reporter les suspens",
                         use_container_width=True,
                         help="Suspens non régularisés → reportés au mois suivant en rouge"):
                if not periode.strip():
                    st.error("⚠️ Saisissez d'abord la **Période**")
                else:
                    cloturer_mois(periode, rb_calc, cp_calc, erb_html_str=erb_html)
                    st.success(f"✅ Mois **{periode}** clôturé. ERB sauvegardé dans l'historique.")
                    st.session_state.page = "dashboard"
                    st.rerun()
    else:
        st.info("ℹ️ Importez d'abord le relevé bancaire et les écritures comptables.")